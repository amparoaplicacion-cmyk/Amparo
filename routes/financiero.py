import os
import smtplib
import ssl
from datetime import date, datetime, timedelta
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

from flask import (Blueprint, flash, redirect, render_template,
                   request, send_file, session, url_for)

from database import get_db
from auth import _cfg_db

financiero_bp = Blueprint('financiero', __name__, url_prefix='/admin/financiero')


# ---------------------------------------------------------------------------
# Guards y contexto
# ---------------------------------------------------------------------------

@financiero_bp.before_request
def verificar_financiero():
    if 'usuario_id' not in session:
        return redirect(url_for('auth.login'))
    if session.get('tipo') != 'admin_financiero':
        flash('No tenés permisos para acceder a esta sección.', 'error')
        return redirect(url_for('admin.dashboard'))
    db = get_db()
    u  = db.execute('SELECT estado FROM usuarios WHERE id=?',
                    (session['usuario_id'],)).fetchone()
    if u and u['estado'] != 'ACTIVA':
        return redirect(url_for('auth.login'))


def _fctx():
    db  = get_db()
    hoy = date.today().isoformat()
    return {
        'nombre':               session.get('nombre', ''),
        'apellido':             session.get('apellido', ''),
        'tipo_usuario':         'admin_financiero',
        'contactos_nuevos':     db.execute(
            "SELECT COUNT(*) AS c FROM contactos WHERE estado='NUEVO'"
        ).fetchone()['c'],
        'conflictos_activos':   db.execute(
            "SELECT COUNT(*) AS c FROM servicios WHERE conflicto=1 AND estado='ACTIVO'"
        ).fetchone()['c'],
        'servicios_activos_hoy': db.execute(
            "SELECT COUNT(*) AS c FROM servicios "
            "WHERE estado IN ('ACEPTADO','ACTIVO') AND fecha_servicio=?", (hoy,)
        ).fetchone()['c'],
    }


# ---------------------------------------------------------------------------
# Helpers compartidos (importables desde otros módulos)
# ---------------------------------------------------------------------------

def registrar_movimiento(db, tipo, descripcion, monto_entrada=0, monto_salida=0,
                         referencia=None, pago_id=None, usuario_id=None):
    """Inserta un movimiento con saldo acumulado running. No hace commit."""
    ultimo = db.execute(
        'SELECT saldo_acumulado FROM movimientos_financieros ORDER BY fecha DESC, id DESC LIMIT 1'
    ).fetchone()
    saldo_ant  = ultimo['saldo_acumulado'] if ultimo else 0
    nuevo_saldo = saldo_ant + (monto_entrada or 0) - (monto_salida or 0)
    print(f'[MOVIMIENTO] Registrando: {tipo} entrada=${monto_entrada or 0} salida={monto_salida or 0} saldo={nuevo_saldo}')
    db.execute(
        '''INSERT INTO movimientos_financieros
           (fecha, tipo, descripcion, monto_entrada, monto_salida,
            saldo_acumulado, referencia, pago_id, usuario_id)
           VALUES (datetime('now', '-3 hours'), ?, ?, ?, ?, ?, ?, ?, ?)''',
        (tipo, descripcion, monto_entrada or 0, monto_salida or 0,
         nuevo_saldo, referencia, pago_id, usuario_id)
    )


def _enviar_excel_email(destinatario, asunto, cuerpo_texto, ruta_archivo, nombre_archivo):
    smtp_host = os.environ.get('SMTP_HOST') or _cfg_db('SMTP_HOST', 'smtp.gmail.com')
    smtp_port = int(os.environ.get('SMTP_PORT') or _cfg_db('SMTP_PORT', '587'))
    smtp_user = os.environ.get('SMTP_USER') or _cfg_db('SMTP_USER', '')
    smtp_pass = os.environ.get('SMTP_PASS') or _cfg_db('SMTP_PASS', '')

    if not smtp_user or not smtp_pass:
        print(f'[AMPARO] Cierre para {destinatario} — SMTP no configurado.')
        return False

    msg = MIMEMultipart()
    msg['Subject'] = asunto
    msg['From']    = smtp_user
    msg['To']      = destinatario
    msg.attach(MIMEText(cuerpo_texto, 'plain', 'utf-8'))

    try:
        with open(ruta_archivo, 'rb') as f:
            part = MIMEBase(
                'application',
                'vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition',
                        f'attachment; filename="{nombre_archivo}"')
        msg.attach(part)
    except Exception as e:
        print(f'[AMPARO] Error adjuntando Excel: {e}')
        return False

    try:
        context = ssl.create_default_context()
        with smtplib.SMTP_SSL(smtp_host, 465, context=context) as server:
            server.login(smtp_user, smtp_pass)
            server.sendmail(smtp_user, destinatario, msg.as_string())
        return True
    except Exception as e:
        print(f'[AMPARO] Error enviando cierre: {e}')
        return False


def generar_cierre_diario(fecha):
    """
    Genera el Excel de movimientos del día indicado y lo envía por email.
    Guarda el archivo en backups/financiero/. Retorna la ruta o None.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print('[AMPARO] openpyxl no instalado — cierre diario no generado.')
        return None

    db = get_db()

    movimientos = db.execute(
        'SELECT * FROM movimientos_financieros WHERE date(fecha) = ? ORDER BY fecha ASC',
        (fecha,)
    ).fetchall()

    saldo_ant_row = db.execute(
        '''SELECT COALESCE(
               (SELECT saldo_acumulado FROM movimientos_financieros
                WHERE date(fecha) < ? ORDER BY fecha DESC LIMIT 1),
           0) AS s''',
        (fecha,)
    ).fetchone()
    saldo_anterior = saldo_ant_row['s'] if saldo_ant_row else 0

    # ── Crear workbook ──────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Movimientos {fecha}'

    # Fila 1: título
    ws.merge_cells('A1:F1')
    ws['A1'] = f'AMPARO — Movimientos del {fecha}'
    ws['A1'].font      = Font(bold=True, size=14, color='1A5276')
    ws['A1'].alignment = Alignment(horizontal='left')

    # Fila 3: saldo inicial
    ws.merge_cells('A3:F3')
    ws['A3'] = f'Saldo al inicio del día: $ {saldo_anterior:,.0f}'
    ws['A3'].font      = Font(bold=True, size=11)
    ws['A3'].alignment = Alignment(horizontal='left')

    # Fila 5: encabezados de columna (sin Referencia)
    encabezados = ['Hora', 'Tipo', 'Descripción', 'Entrada $', 'Salida $', 'Saldo $']
    for col, enc in enumerate(encabezados, 1):
        cell = ws.cell(row=5, column=col, value=enc)
        cell.font      = Font(bold=True, color='FFFFFF')
        cell.fill      = PatternFill(fill_type='solid', fgColor='1A5276')
        cell.alignment = Alignment(horizontal='center')

    colores = {
        'COBRO':                  'D5F5E3',
        'PAGO_PRESTADOR':         'FDEBD0',
        'TRANSFERENCIA_PERSONAL': 'D6EAF8',
        'PENALIDAD':              'FADBD8',
    }

    total_entradas = 0
    total_salidas  = 0
    saldo_final    = saldo_anterior

    for fila, mov in enumerate(movimientos, 6):
        try:
            hora = datetime.fromisoformat(str(mov['fecha'])).strftime('%H:%M:%S')
        except Exception:
            hora = str(mov['fecha'])[-8:]
        color   = colores.get(mov['tipo'], 'FFFFFF')
        entrada = mov['monto_entrada'] or 0
        salida  = mov['monto_salida']  or 0
        saldo_final = mov['saldo_acumulado']
        datos = [hora, mov['tipo'], mov['descripcion'],
                 entrada if entrada else '',
                 salida  if salida  else '',
                 mov['saldo_acumulado']]
        for col, dato in enumerate(datos, 1):
            cell = ws.cell(row=fila, column=col, value=dato)
            cell.fill = PatternFill(fill_type='solid', fgColor=color)
        total_entradas += entrada
        total_salidas  += salida

    # Fila de totales
    fila_tot = len(movimientos) + 6
    ws.cell(row=fila_tot, column=1, value='TOTALES DEL DÍA').font = Font(bold=True)
    ws.cell(row=fila_tot, column=4, value=total_entradas).font    = Font(bold=True)
    ws.cell(row=fila_tot, column=5, value=total_salidas).font     = Font(bold=True)
    cell_sf = ws.cell(row=fila_tot, column=6,
                      value=f'{saldo_final:,.0f} — Saldo Final')
    cell_sf.font = Font(bold=True, color='1A5276', size=12)
    cell_sf.fill = PatternFill(fill_type='solid', fgColor='D6EAF8')

    # Anchos de columna
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 22

    # ── Guardar ──────────────────────────────────────────────────────────────
    carpeta = os.path.join('backups', 'financiero')
    os.makedirs(carpeta, exist_ok=True)
    nombre = f'movimientos_{fecha}.xlsx'
    ruta   = os.path.join(carpeta, nombre)
    wb.save(ruta)

    # ── Enviar email ─────────────────────────────────────────────────────────
    email_dest = _cfg_db('admin_financiero_email', '')
    if email_dest:
        saldo_cierre = saldo_anterior + total_entradas - total_salidas
        asunto = f'AMPARO — Movimientos del {fecha}'
        cuerpo = (
            f'Resumen de movimientos financieros del {fecha}.\n\n'
            f'Total entradas:  $ {total_entradas:,.2f}\n'
            f'Total salidas:   $ {total_salidas:,.2f}\n'
            f'Saldo al cierre: $ {saldo_cierre:,.2f}\n\n'
            'Se adjunta el detalle completo en Excel.\n\n'
            'El equipo de AMPARO'
        )
        _enviar_excel_email(email_dest, asunto, cuerpo, ruta, nombre)

    return ruta


def verificar_cierre_diario():
    """
    Genera cierres diarios faltantes (hasta ayer).
    Llamar tras el login del admin_financiero.
    """
    db         = get_db()
    hoy        = date.today()
    ayer       = hoy - timedelta(days=1)
    ultimo_str = _cfg_db('ultimo_cierre_diario', '')

    if not ultimo_str or ultimo_str < ayer.isoformat():
        if not ultimo_str:
            fecha_inicio = ayer
        else:
            fecha_inicio = date.fromisoformat(ultimo_str) + timedelta(days=1)

        fecha_actual = fecha_inicio
        while fecha_actual <= ayer:
            try:
                generar_cierre_diario(fecha_actual.isoformat())
            except Exception as e:
                print(f'[AMPARO] Error generando cierre {fecha_actual}: {e}')
            fecha_actual += timedelta(days=1)

        db.execute("UPDATE configuracion SET valor=? WHERE clave=?",
                   (ayer.isoformat(), 'ultimo_cierre_diario'))
        db.commit()


# ---------------------------------------------------------------------------
# Ruta: Movimientos
# ---------------------------------------------------------------------------

@financiero_bp.route('/movimientos')
def movimientos():
    db          = get_db()
    fecha_desde = request.args.get('desde', '')
    fecha_hasta = request.args.get('hasta', '')

    query  = 'SELECT * FROM movimientos_financieros WHERE 1=1'
    params = []
    if fecha_desde:
        query  += ' AND date(fecha) >= ?'
        params.append(fecha_desde)
    if fecha_hasta:
        query  += ' AND date(fecha) <= ?'
        params.append(fecha_hasta)
    query += ' ORDER BY fecha DESC'

    movs = db.execute(query, params).fetchall()

    return render_template(
        'admin/financiero/movimientos.html',
        seccion_activa='fin_movimientos',
        movs=movs,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        **_fctx()
    )


@financiero_bp.route('/movimientos/registrar', methods=['POST'])
def movimientos_registrar():
    monto_str   = request.form.get('monto', '').replace(',', '.').strip()
    descripcion = request.form.get('descripcion', '').strip()
    referencia  = request.form.get('referencia', '').strip() or None

    try:
        monto = float(monto_str)
        if monto <= 0:
            raise ValueError
    except (ValueError, TypeError):
        flash('El monto debe ser un número mayor a cero.', 'error')
        return redirect(url_for('financiero.movimientos'))

    if not descripcion:
        flash('La descripción es obligatoria.', 'error')
        return redirect(url_for('financiero.movimientos'))

    db = get_db()
    registrar_movimiento(
        db, 'TRANSFERENCIA_PERSONAL', descripcion,
        monto_salida=monto, referencia=referencia,
        usuario_id=session.get('usuario_id')
    )
    db.commit()
    flash(f'Transferencia de ${monto:,.2f} registrada correctamente.', 'success')
    return redirect(url_for('financiero.movimientos'))


# ---------------------------------------------------------------------------
# Ruta: Saldo y Acumulados
# ---------------------------------------------------------------------------

@financiero_bp.route('/saldo')
def saldo():
    db = get_db()

    ultimo = db.execute(
        'SELECT saldo_acumulado FROM movimientos_financieros ORDER BY id DESC LIMIT 1'
    ).fetchone()
    saldo_actual = ultimo['saldo_acumulado'] if ultimo else 0

    totales = db.execute(
        '''SELECT tipo,
                  SUM(monto_entrada) AS total_entrada,
                  SUM(monto_salida)  AS total_salida,
                  COUNT(*)           AS cantidad
           FROM movimientos_financieros
           GROUP BY tipo
           ORDER BY tipo'''
    ).fetchall()

    por_mes = db.execute(
        '''SELECT strftime('%Y-%m', fecha) AS mes,
                  SUM(monto_entrada) AS entradas,
                  SUM(monto_salida)  AS salidas
           FROM movimientos_financieros
           GROUP BY mes
           ORDER BY mes DESC
           LIMIT 12'''
    ).fetchall()

    return render_template(
        'admin/financiero/saldo.html',
        seccion_activa='fin_saldo',
        saldo_actual=saldo_actual,
        totales=totales,
        por_mes=por_mes,
        **_fctx()
    )


# ---------------------------------------------------------------------------
# Ruta: Historial de Cierres
# ---------------------------------------------------------------------------

@financiero_bp.route('/cierres')
def cierres():
    carpeta  = os.path.join('backups', 'financiero')
    archivos = []
    if os.path.isdir(carpeta):
        for f in sorted(os.listdir(carpeta), reverse=True):
            if f.startswith('movimientos_') and f.endswith('.xlsx'):
                ruta      = os.path.join(carpeta, f)
                fecha_str = f[len('movimientos_'):-len('.xlsx')]
                archivos.append({
                    'nombre':  f,
                    'fecha':   fecha_str,
                    'size_kb': round(os.path.getsize(ruta) / 1024, 1),
                })

    return render_template(
        'admin/financiero/cierres.html',
        seccion_activa='fin_cierres',
        archivos=archivos,
        **_fctx()
    )


@financiero_bp.route('/cierres/<fecha>/descargar')
def cierre_descargar(fecha):
    ruta = os.path.join('backups', 'financiero', f'movimientos_{fecha}.xlsx')
    if not os.path.isfile(ruta):
        flash('Archivo no encontrado.', 'error')
        return redirect(url_for('financiero.cierres'))
    return send_file(
        ruta,
        as_attachment=True,
        download_name=f'movimientos_{fecha}.xlsx',
        mimetype=(
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    )


@financiero_bp.route('/cierres/<fecha>/regenerar', methods=['POST'])
def cierre_regenerar(fecha):
    try:
        ruta = generar_cierre_diario(fecha)
        if ruta:
            flash(f'Cierre del {fecha} regenerado correctamente.', 'success')
        else:
            flash('No se pudo regenerar (openpyxl no instalado).', 'error')
    except Exception as e:
        flash(f'Error al regenerar: {e}', 'error')
    return redirect(url_for('financiero.cierres'))


# ---------------------------------------------------------------------------
# Ruta: Configuración Financiera
# ---------------------------------------------------------------------------

@financiero_bp.route('/configuracion', methods=['GET', 'POST'])
def configuracion():
    db = get_db()

    if request.method == 'POST':
        email_fin    = request.form.get('admin_financiero_email', '').strip()
        hora_cierre  = request.form.get('cierre_diario_hora', '00:00').strip()
        anios_str    = request.form.get('movimientos_backup_anios', '10').strip()
        try:
            dias_ret = int(float(anios_str)) * 365
        except ValueError:
            dias_ret = 3650

        for clave, valor in [
            ('admin_financiero_email',  email_fin),
            ('cierre_diario_hora',      hora_cierre),
            ('movimientos_backup_dias', str(dias_ret)),
        ]:
            db.execute('UPDATE configuracion SET valor=? WHERE clave=?', (valor, clave))
        db.commit()
        flash('Configuración financiera guardada.', 'success')
        return redirect(url_for('financiero.configuracion'))

    cfg = {r['clave']: r['valor'] for r in db.execute(
        "SELECT clave, valor FROM configuracion "
        "WHERE clave IN ('admin_financiero_email','cierre_diario_hora','movimientos_backup_dias')"
    ).fetchall()}
    anios_ret = round(int(cfg.get('movimientos_backup_dias', 3650)) / 365)

    return render_template(
        'admin/financiero/configuracion.html',
        seccion_activa='fin_configuracion',
        cfg=cfg,
        anios_ret=anios_ret,
        **_fctx()
    )
