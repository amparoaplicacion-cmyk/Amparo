import os
import secrets
import string
from datetime import date, datetime, timedelta

from flask import (Blueprint, flash, redirect, render_template,
                   request, session, url_for)
from werkzeug.security import generate_password_hash
from werkzeug.utils import secure_filename

from database import get_db, ahora_argentina
from auth import enviar_email, _cfg_db

admin_bp = Blueprint('admin', __name__, url_prefix='/admin')


def _calcular_horas(hora_inicio, hora_fin):
    """Calcula horas exactas entre dos horarios HH:MM (acepta también HH:MM:SS)."""
    hi = datetime.strptime(hora_inicio[:5], '%H:%M')
    hf = datetime.strptime(hora_fin[:5], '%H:%M')
    diferencia = hf - hi
    return diferencia.seconds / 3600

_BASE_DIR           = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
UPLOAD_FOLDER       = os.path.join(_BASE_DIR, 'static', 'uploads', 'prestadores')
ALLOWED_EXTS        = {'jpg', 'jpeg', 'png', 'webp'}
ANTECEDENTES_FOLDER = os.path.join(_BASE_DIR, 'static', 'docs', 'antecedentes')


# ---------------------------------------------------------------------------
# Guard de sesión para todo el blueprint
# ---------------------------------------------------------------------------

@admin_bp.before_request
def verificar_admin():
    if 'usuario_id' not in session:
        return redirect(url_for('auth.login'))
    if session.get('tipo') not in ('admin', 'admin_financiero'):
        return redirect(url_for('auth.login'))
    db = get_db()
    u = db.execute("SELECT estado FROM usuarios WHERE id=?", (session['usuario_id'],)).fetchone()
    if u and u['estado'] == 'VENCIDA':
        session['cambio_requerido'] = True
        return redirect(url_for('auth.cambiar_password'))


def _ctx():
    db  = get_db()
    hoy = date.today().isoformat()
    contactos_nuevos = db.execute(
        "SELECT COUNT(*) as c FROM contactos WHERE estado='NUEVO'"
    ).fetchone()['c']
    conflictos_activos = db.execute(
        "SELECT COUNT(*) as c FROM servicios WHERE conflicto=1 AND estado='ACTIVO'"
    ).fetchone()['c']
    servicios_activos_hoy = db.execute(
        "SELECT COUNT(*) as c FROM servicios WHERE estado IN ('ACEPTADO','ACTIVO') AND fecha_servicio=?",
        (hoy,)
    ).fetchone()['c']
    return {
        'nombre':               session.get('nombre', ''),
        'apellido':             session.get('apellido', ''),
        'tipo_usuario':         session.get('tipo', ''),
        'contactos_nuevos':     contactos_nuevos,
        'conflictos_activos':   conflictos_activos,
        'servicios_activos_hoy': servicios_activos_hoy,
    }


# ---------------------------------------------------------------------------
# Helpers compartidos
# ---------------------------------------------------------------------------

def generar_password_temporal():
    chars    = string.ascii_letters + string.digits
    specials = '!@#$%^&*'
    pwd = [
        secrets.choice(string.ascii_uppercase),
        secrets.choice(specials),
        secrets.choice(string.ascii_lowercase),
        secrets.choice(string.digits),
    ]
    for _ in range(6):
        pwd.append(secrets.choice(chars))
    secrets.SystemRandom().shuffle(pwd)
    return ''.join(pwd)


def _guardar_foto(file, prefix):
    if not file or not file.filename:
        return None
    ext = file.filename.rsplit('.', 1)[-1].lower()
    if ext not in ALLOWED_EXTS:
        return None
    filename = f'{prefix}.{ext}'
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
    file.save(os.path.join(UPLOAD_FOLDER, filename))
    return filename


def _notificar(db, usuario_id, tipo, titulo, mensaje=None):
    db.execute(
        'INSERT INTO notificaciones (usuario_id, tipo, titulo, mensaje) VALUES (?, ?, ?, ?)',
        (usuario_id, tipo, titulo, mensaje)
    )


def _get_prestador(db, pid):
    return db.execute(
        '''SELECT p.*, u.nombre, u.apellido, u.email, u.telefono, u.fecha_alta,
                  c.nombre AS categoria_nombre,
                  z.nombre AS zona_nombre
           FROM prestadores p
           JOIN usuarios u ON p.usuario_id = u.id
           LEFT JOIN categorias c ON p.categoria_id = c.id
           LEFT JOIN zonas z ON p.zona_id = z.id
           WHERE p.id = ?''',
        (pid,)
    ).fetchone()


# ---------------------------------------------------------------------------
# Dashboard principal
# ---------------------------------------------------------------------------

@admin_bp.route('/')
@admin_bp.route('/dashboard')
def dashboard():
    db = get_db()

    usuarios_activos = db.execute(
        "SELECT COUNT(*) FROM usuarios WHERE estado = 'ACTIVA'"
    ).fetchone()[0]

    prestadores_pendientes = db.execute(
        "SELECT COUNT(*) FROM prestadores WHERE estado_perfil = 'EN_REVISION'"
    ).fetchone()[0]

    hoy = date.today().isoformat()

    servicios_hoy = db.execute(
        "SELECT COUNT(*) FROM servicios WHERE estado IN ('ACEPTADO','ACTIVO') AND fecha_servicio = ?",
        (hoy,)
    ).fetchone()[0]

    reclamos_abiertos = db.execute(
        "SELECT COUNT(*) FROM reclamos WHERE estado = 'ABIERTO'"
    ).fetchone()[0]

    notificaciones = db.execute(
        '''SELECT n.fecha, n.tipo, n.titulo, n.mensaje,
                  u.nombre || ' ' || u.apellido AS destinatario
           FROM notificaciones n
           JOIN usuarios u ON n.usuario_id = u.id
           ORDER BY n.fecha DESC LIMIT 5'''
    ).fetchall()

    servicios_hoy_lista = db.execute(
        '''SELECT s.id, s.hora_inicio, s.hora_fin, s.estado,
                  s.prestador_confirmo_llegada, s.prestador_confirmo_fin,
                  s.solicitante_confirmo_fin, s.conflicto,
                  pu.nombre || ' ' || pu.apellido AS prestador_nombre,
                  fu.nombre || ' ' || fu.apellido AS solicitante_nombre,
                  c.nombre AS categoria_nombre
           FROM servicios s
           JOIN prestadores p ON s.prestador_id = p.id
           JOIN usuarios pu   ON p.usuario_id   = pu.id
           JOIN solicitantes  f   ON s.solicitante_id   = f.id
           JOIN usuarios fu   ON f.usuario_id   = fu.id
           LEFT JOIN categorias c ON p.categoria_id = c.id
           WHERE s.fecha_servicio = ? AND s.estado IN ('ACEPTADO','ACTIVO')
           ORDER BY s.hora_inicio''',
        (hoy,)
    ).fetchall()

    # Servicios donde el prestador confirmó fin pero el solicitante aún no
    confirmacion_pendiente = db.execute(
        '''SELECT s.id, s.fecha_servicio, s.hora_inicio, s.hora_fin,
                  s.fecha_confirmacion_prestador,
                  pu.nombre || ' ' || pu.apellido AS prestador_nombre,
                  fu.nombre || ' ' || fu.apellido AS solicitante_nombre,
                  fu_u.id AS solicitante_usuario_id,
                  c.nombre AS categoria_nombre
           FROM servicios s
           JOIN prestadores p ON s.prestador_id = p.id
           JOIN usuarios pu   ON p.usuario_id   = pu.id
           JOIN solicitantes  f   ON s.solicitante_id   = f.id
           JOIN usuarios fu   ON f.usuario_id   = fu.id
           JOIN usuarios fu_u ON fu_u.id = f.usuario_id
           LEFT JOIN categorias c ON p.categoria_id = c.id
           WHERE s.prestador_confirmo_fin=1 AND s.solicitante_confirmo_fin=0
             AND s.estado IN ('ACEPTADO','ACTIVO')
           ORDER BY s.fecha_confirmacion_prestador ASC''',
    ).fetchall()

    return render_template(
        'admin/dashboard.html',
        seccion_activa='dashboard',
        usuarios_activos=usuarios_activos,
        prestadores_pendientes=prestadores_pendientes,
        servicios_hoy=servicios_hoy,
        reclamos_abiertos=reclamos_abiertos,
        notificaciones=notificaciones,
        servicios_hoy_lista=servicios_hoy_lista,
        confirmacion_pendiente=confirmacion_pendiente,
        **_ctx()
    )


# ---------------------------------------------------------------------------
# Confirmación manual de finalización
# ---------------------------------------------------------------------------

@admin_bp.route('/servicios/<int:sid>/recordatorio-confirmacion', methods=['POST'])
def recordatorio_confirmacion(sid):
    db = get_db()
    s = db.execute(
        '''SELECT s.*, fu.email AS sol_email, fu.nombre AS sol_nombre
           FROM servicios s
           JOIN solicitantes sol ON sol.id = s.solicitante_id
           JOIN usuarios fu ON fu.id = sol.usuario_id
           WHERE s.id=? AND s.prestador_confirmo_fin=1 AND s.solicitante_confirmo_fin=0''',
        (sid,)
    ).fetchone()
    if not s:
        flash('No se puede enviar el recordatorio.', 'error')
        return redirect(url_for('admin.dashboard'))

    app_url = _cfg_db('app_url', 'http://127.0.0.1:5000')
    asunto  = 'Recordatorio: confirmá la finalización del servicio — AMPARO'
    cuerpo  = (
        f'Hola {s["sol_nombre"]},\n\n'
        f'El prestador confirmó que el servicio del {s["fecha_servicio"]} '
        f'({s["hora_inicio"]} a {s["hora_fin"]}) fue completado.\n\n'
        f'Por favor ingresá a la app y confirmá la finalización:\n'
        f'{app_url}/login'
    )
    enviar_email(s['sol_email'], asunto, cuerpo)
    flash(f'Recordatorio enviado a {s["sol_email"]}.', 'success')
    return redirect(url_for('admin.dashboard'))


@admin_bp.route('/servicios/<int:sid>/confirmar-fin-manual', methods=['POST'])
def confirmar_fin_manual(sid):
    db    = get_db()
    ahora = ahora_argentina()
    s = db.execute(
        "SELECT * FROM servicios WHERE id=? AND prestador_confirmo_fin=1 AND solicitante_confirmo_fin=0",
        (sid,)
    ).fetchone()
    if not s:
        flash('No se puede confirmar este servicio.', 'error')
        return redirect(url_for('admin.dashboard'))

    db.execute(
        """UPDATE servicios SET solicitante_confirmo_fin=1,
           fecha_confirmacion_solicitante=?, estado='FINALIZADO', fecha_finalizacion=?
           WHERE id=?""",
        (ahora, ahora, sid)
    )

    # Crear pago si no existe
    pago_existente = db.execute(
        "SELECT id FROM pagos WHERE servicio_id=?", (sid,)
    ).fetchone()
    if not pago_existente:
        monto = s['monto_estimado'] or s['monto_acordado'] or 0
        db.execute(
            """INSERT INTO pagos (servicio_id, solicitante_id, prestador_id, tipo_pago,
               monto_bruto, comision_pct, comision_monto, monto_neto,
               estado, metodo_pago, fecha_pago, fecha_liquidacion)
               VALUES (?,?,?,'servicio',?,0,0,?,'LIQUIDADO','admin_manual',?,?)""",
            (sid, s['solicitante_id'], s['prestador_id'],
             monto, monto, ahora, ahora)
        )

    _notificar(db, db.execute(
        "SELECT usuario_id FROM solicitantes WHERE id=?", (s['solicitante_id'],)
    ).fetchone()['usuario_id'],
    'servicio_finalizado', 'Servicio confirmado por AMPARO',
    f'El equipo de AMPARO confirmó la finalización del servicio del {s["fecha_servicio"]}.')

    db.commit()
    flash(f'Servicio #{sid} confirmado manualmente.', 'success')
    return redirect(url_for('admin.dashboard'))


# ---------------------------------------------------------------------------
# Gestión de Usuarios
# ---------------------------------------------------------------------------

@admin_bp.route('/usuarios')
def usuarios():
    busqueda = request.args.get('busqueda', '').strip()
    tipo     = request.args.get('tipo', '')
    estado   = request.args.get('estado', '')

    db     = get_db()
    query  = 'SELECT * FROM usuarios WHERE 1=1'
    params = []

    if busqueda:
        query += ' AND (nombre LIKE ? OR apellido LIKE ? OR email LIKE ?)'
        params += [f'%{busqueda}%', f'%{busqueda}%', f'%{busqueda}%']
    if tipo:
        query += ' AND tipo_usuario = ?'
        params.append(tipo)
    if estado:
        query += ' AND estado = ?'
        params.append(estado)

    query += ' ORDER BY fecha_alta DESC'
    lista  = db.execute(query, params).fetchall()

    return render_template(
        'admin/usuarios/listado.html',
        seccion_activa='usuarios',
        usuarios=lista,
        busqueda=busqueda,
        tipo=tipo,
        estado=estado,
        **_ctx()
    )


@admin_bp.route('/usuarios/<int:uid>')
def usuario_detalle(uid):
    db      = get_db()
    usuario = db.execute('SELECT * FROM usuarios WHERE id = ?', (uid,)).fetchone()
    if not usuario:
        flash('Usuario no encontrado.', 'error')
        return redirect(url_for('admin.usuarios'))
    return render_template(
        'admin/usuarios/detalle.html',
        seccion_activa='usuarios',
        usuario=usuario,
        **_ctx()
    )


@admin_bp.route('/usuarios/nuevo', methods=['GET', 'POST'])
def usuario_nuevo():
    if request.method == 'POST':
        nombre   = request.form.get('nombre', '').strip()
        apellido = request.form.get('apellido', '').strip()
        email    = request.form.get('email', '').strip().lower()
        telefono = request.form.get('telefono', '').strip()
        tipo     = request.form.get('tipo_usuario', '')
        password = request.form.get('password_generada', '')

        errores = []
        if not nombre:   errores.append('El nombre es obligatorio.')
        if not apellido: errores.append('El apellido es obligatorio.')
        if not email:    errores.append('El email es obligatorio.')
        if not tipo:     errores.append('El tipo de usuario es obligatorio.')

        if not errores:
            db = get_db()
            if db.execute('SELECT id FROM usuarios WHERE email = ?', (email,)).fetchone():
                errores.append('Ya existe un usuario con ese email.')
            if tipo == 'admin_financiero':
                if session.get('tipo') != 'admin_financiero':
                    errores.append('Solo el administrador financiero puede crear usuarios de este tipo.')
                elif db.execute(
                    "SELECT id FROM usuarios WHERE tipo_usuario='admin_financiero'",
                ).fetchone():
                    errores.append(
                        'Ya existe un administrador financiero en el sistema. '
                        'Solo puede haber uno.'
                    )

        if errores:
            for e in errores:
                flash(e, 'error')
            return render_template(
                'admin/usuarios/nuevo.html',
                seccion_activa='usuarios',
                password_generada=password,
                form=request.form,
                **_ctx()
            )

        db   = get_db()
        hoy  = date.today().isoformat()
        venc = (date.today() + timedelta(days=90)).isoformat()
        db.execute(
            '''INSERT INTO usuarios
               (nombre, apellido, email, password_hash, tipo_usuario,
                estado, intentos_fallidos, fecha_alta, fecha_password,
                fecha_vencimiento, telefono)
               VALUES (?, ?, ?, ?, ?, 'VENCIDA', 0, ?, ?, ?, ?)''',
            (nombre, apellido, email, generate_password_hash(password),
             tipo, datetime.now().isoformat(), hoy, venc, telefono or None)
        )
        db.commit()
        flash(
            f'Usuario {nombre} {apellido} creado correctamente. '
            f'Contraseña temporal: {password}',
            'success'
        )
        return redirect(url_for('admin.usuarios'))

    return render_template(
        'admin/usuarios/nuevo.html',
        seccion_activa='usuarios',
        password_generada=generar_password_temporal(),
        form={},
        **_ctx()
    )


@admin_bp.route('/usuarios/<int:uid>/editar', methods=['GET', 'POST'])
def usuario_editar(uid):
    db      = get_db()
    usuario = db.execute('SELECT * FROM usuarios WHERE id = ?', (uid,)).fetchone()
    if not usuario:
        flash('Usuario no encontrado.', 'error')
        return redirect(url_for('admin.usuarios'))

    if request.method == 'POST':
        nombre   = request.form.get('nombre', '').strip()
        apellido = request.form.get('apellido', '').strip()
        telefono = request.form.get('telefono', '').strip()
        tipo     = request.form.get('tipo_usuario', '')
        estado   = request.form.get('estado', '')

        errores = []
        if not nombre:   errores.append('El nombre es obligatorio.')
        if not apellido: errores.append('El apellido es obligatorio.')
        if not tipo:     errores.append('El tipo de usuario es obligatorio.')
        if not estado:   errores.append('El estado es obligatorio.')

        if errores:
            for e in errores:
                flash(e, 'error')
            return render_template(
                'admin/usuarios/editar.html',
                seccion_activa='usuarios',
                usuario=usuario,
                **_ctx()
            )

        db.execute(
            '''UPDATE usuarios
               SET nombre = ?, apellido = ?, telefono = ?,
                   tipo_usuario = ?, estado = ?
               WHERE id = ?''',
            (nombre, apellido, telefono or None, tipo, estado, uid)
        )
        db.commit()
        flash('Usuario actualizado correctamente.', 'success')
        return redirect(url_for('admin.usuario_detalle', uid=uid))

    return render_template(
        'admin/usuarios/editar.html',
        seccion_activa='usuarios',
        usuario=usuario,
        **_ctx()
    )


@admin_bp.route('/usuarios/<int:uid>/bloquear', methods=['POST'])
def usuario_bloquear(uid):
    db = get_db()
    db.execute(
        "UPDATE usuarios SET estado = 'BLOQUEADA', fecha_bloqueo = ? WHERE id = ?",
        (datetime.now().isoformat(), uid)
    )
    db.commit()
    flash('Cuenta bloqueada correctamente.', 'success')
    return redirect(url_for('admin.usuario_detalle', uid=uid))


@admin_bp.route('/usuarios/<int:uid>/desbloquear', methods=['POST'])
def usuario_desbloquear(uid):
    db = get_db()
    db.execute(
        '''UPDATE usuarios
           SET estado = 'ACTIVA', intentos_fallidos = 0,
               fecha_bloqueo = NULL, token_desbloqueo = NULL, token_expira = NULL
           WHERE id = ?''',
        (uid,)
    )
    db.commit()
    flash('Cuenta desbloqueada correctamente.', 'success')
    return redirect(url_for('admin.usuario_detalle', uid=uid))


@admin_bp.route('/usuarios/<int:uid>/resetear_password', methods=['POST'])
def usuario_resetear_password(uid):
    nueva = generar_password_temporal()
    hoy   = date.today().isoformat()
    venc  = (date.today() + timedelta(days=90)).isoformat()
    db    = get_db()
    db.execute(
        '''UPDATE usuarios
           SET password_hash = ?, fecha_password = ?, fecha_vencimiento = ?,
               estado = 'VENCIDA', intentos_fallidos = 0
           WHERE id = ?''',
        (generate_password_hash(nueva), hoy, venc, uid)
    )
    db.commit()
    u = db.execute('SELECT nombre, email, tipo_usuario FROM usuarios WHERE id=?', (uid,)).fetchone()
    if u:
        _base_url = _cfg_db('app_url', 'http://127.0.0.1:5000')
        _tipo_u   = u['tipo_usuario'] or ''
        _link_app = _base_url + '/login'
        asunto = _cfg_db('mail_contrasena_temp_asunto', 'Tu contraseña temporal — AMPARO')
        cuerpo = _cfg_db('mail_contrasena_temp_cuerpo',
            'Hola {nombre},\n\nTu contraseña fue reseteada por el administrador.\n\n'
            'Tu contraseña temporal es:\n\n{contrasena_temporal}\n\n'
            'Deberás cambiarla al ingresar.\n\n{link_app}')
        cuerpo = (cuerpo
            .replace('{nombre}', u['nombre'])
            .replace('{contrasena_temporal}', nueva)
            .replace('{empresa_email}', _cfg_db('empresa_email', ''))
            .replace('{link_app}', _link_app)
        )
        enviar_email(u['email'], asunto, cuerpo)
    flash(f'Contraseña reseteada. Nueva contraseña temporal: {nueva}', 'success')
    return redirect(url_for('admin.usuario_detalle', uid=uid))


@admin_bp.route('/usuarios/<int:uid>/desactivar', methods=['POST'])
def usuario_desactivar(uid):
    db = get_db()
    db.execute("UPDATE usuarios SET estado = 'INACTIVA' WHERE id = ?", (uid,))
    db.commit()
    flash('Cuenta desactivada correctamente.', 'success')
    return redirect(url_for('admin.usuario_detalle', uid=uid))


@admin_bp.route('/usuarios/<int:uid>/eliminar', methods=['POST'])
def usuario_eliminar(uid):
    """Elimina permanentemente un usuario y todos sus datos relacionados."""
    db = get_db()
    u = db.execute('SELECT email, tipo_usuario FROM usuarios WHERE id=?', (uid,)).fetchone()
    if not u:
        flash('Usuario no encontrado.', 'error')
        return redirect(url_for('admin.usuarios'))
    if u['tipo_usuario'] in ('admin', 'admin_financiero'):
        flash('No se puede eliminar un usuario administrador.', 'error')
        return redirect(url_for('admin.usuario_detalle', uid=uid))

    # Notificar al admin antes de borrar los datos
    tipo_label = {'prestador': 'Prestador', 'solicitante': 'Solicitante'}.get(u['tipo_usuario'], u['tipo_usuario'])
    _notificar(db, session['usuario_id'], 'BAJA',
               f'{tipo_label} eliminado: {u["email"]}',
               f'{tipo_label} eliminado permanentemente del sistema.')

    # Eliminar datos relacionados según tipo
    sol = db.execute('SELECT id FROM solicitantes WHERE usuario_id=?', (uid,)).fetchone()
    if sol:
        db.execute('DELETE FROM notificaciones WHERE usuario_id=?', (uid,))
        db.execute('DELETE FROM solicitantes WHERE usuario_id=?', (uid,))

    pre = db.execute('SELECT id FROM prestadores WHERE usuario_id=?', (uid,)).fetchone()
    if pre:
        db.execute('DELETE FROM disponibilidad WHERE prestador_id=?', (pre['id'],))
        db.execute('DELETE FROM notificaciones WHERE usuario_id=?', (uid,))
        db.execute('DELETE FROM prestadores WHERE usuario_id=?', (uid,))

    db.execute('DELETE FROM notificaciones WHERE usuario_id=?', (uid,))
    db.execute('DELETE FROM usuarios WHERE id=?', (uid,))
    db.commit()
    flash(f'Usuario {u["email"]} eliminado permanentemente.', 'success')
    return redirect(url_for('admin.usuarios'))


@admin_bp.route('/diagnostico-db')
def diagnostico_db():
    """Muestra la ruta de la base de datos activa y todos los usuarios — solo admins."""
    from database import DATABASE
    db = get_db()
    usuarios = db.execute(
        'SELECT id, nombre, apellido, email, tipo_usuario, estado, fecha_alta FROM usuarios ORDER BY id'
    ).fetchall()
    import os
    existe = os.path.isfile(DATABASE)
    tamano = os.path.getsize(DATABASE) if existe else 0
    lineas = [
        f'<h2>Diagnóstico de Base de Datos</h2>',
        f'<p><b>Ruta:</b> {DATABASE}</p>',
        f'<p><b>Existe:</b> {"Sí" if existe else "NO"}</p>',
        f'<p><b>Tamaño:</b> {tamano} bytes</p>',
        f'<p><b>Total usuarios:</b> {len(usuarios)}</p>',
        f'<table border="1" cellpadding="4">',
        f'<tr><th>ID</th><th>Email</th><th>Nombre</th><th>Tipo</th><th>Estado</th><th>Alta</th></tr>',
    ]
    for u in usuarios:
        lineas.append(
            f'<tr><td>{u["id"]}</td><td>{u["email"]}</td>'
            f'<td>{u["nombre"]} {u["apellido"]}</td>'
            f'<td>{u["tipo_usuario"]}</td><td>{u["estado"]}</td>'
            f'<td>{u["fecha_alta"] or "—"}</td></tr>'
        )
    lineas.append('</table>')
    lineas.append('<br><a href="/admin/usuarios">← Volver a usuarios</a>')
    return '\n'.join(lineas), 200


# ---------------------------------------------------------------------------
# Gestión de Prestadores — Listado
# ---------------------------------------------------------------------------

@admin_bp.route('/prestadores')
def prestadores():
    busqueda      = request.args.get('busqueda', '').strip()
    categoria_id  = request.args.get('categoria_id', '')
    estado_perfil = request.args.get('estado_perfil', '')
    zona_id       = request.args.get('zona_id', '')

    db     = get_db()
    query  = '''
        SELECT p.*, u.nombre, u.apellido, u.email, u.fecha_alta,
               c.nombre AS categoria_nombre,
               z.nombre AS zona_nombre
        FROM prestadores p
        JOIN usuarios u ON p.usuario_id = u.id
        LEFT JOIN categorias c ON p.categoria_id = c.id
        LEFT JOIN zonas z ON p.zona_id = z.id
        WHERE 1=1
    '''
    params = []

    if busqueda:
        query += ' AND (u.nombre LIKE ? OR u.apellido LIKE ? OR u.email LIKE ?)'
        params += [f'%{busqueda}%', f'%{busqueda}%', f'%{busqueda}%']
    if categoria_id:
        query += ' AND p.categoria_id = ?'
        params.append(int(categoria_id))
    if estado_perfil:
        query += ' AND p.estado_perfil = ?'
        params.append(estado_perfil)
    if zona_id:
        query += ' AND p.zona_id = ?'
        params.append(int(zona_id))

    query += ' ORDER BY u.fecha_alta DESC'

    lista      = db.execute(query, params).fetchall()
    categorias = db.execute('SELECT * FROM categorias WHERE activa=1 ORDER BY nombre').fetchall()
    zonas      = db.execute('SELECT * FROM zonas WHERE activa=1 ORDER BY nombre').fetchall()

    hoy_date  = date.today()
    hoy_iso   = hoy_date.isoformat()
    alert_iso = (hoy_date + timedelta(days=30)).isoformat()

    return render_template(
        'admin/prestadores/listado.html',
        seccion_activa='prestadores',
        prestadores=lista,
        categorias=categorias,
        zonas=zonas,
        busqueda=busqueda,
        categoria_id=categoria_id,
        estado_perfil=estado_perfil,
        zona_id=zona_id,
        hoy_iso=hoy_iso,
        alert_iso=alert_iso,
        **_ctx()
    )


# ---------------------------------------------------------------------------
# Gestión de Prestadores — Detalle
# ---------------------------------------------------------------------------

@admin_bp.route('/prestadores/<int:pid>')
def prestador_detalle(pid):
    db = get_db()
    p  = _get_prestador(db, pid)
    if not p:
        flash('Prestador no encontrado.', 'error')
        return redirect(url_for('admin.prestadores'))

    servicios = db.execute(
        '''SELECT s.fecha_servicio, s.monto_acordado, s.estado,
                  u.nombre || ' ' || u.apellido AS solicitante_nombre,
                  cal.puntaje
           FROM servicios s
           JOIN solicitantes f   ON s.solicitante_id   = f.id
           JOIN usuarios u   ON f.usuario_id   = u.id
           LEFT JOIN calificaciones cal ON s.id = cal.servicio_id
           WHERE s.prestador_id = ?
           ORDER BY s.fecha_servicio DESC
           LIMIT 10''',
        (pid,)
    ).fetchall()

    promedio = db.execute(
        'SELECT AVG(puntaje) FROM calificaciones WHERE prestador_id = ?', (pid,)
    ).fetchone()[0]

    calificaciones_list = db.execute(
        '''SELECT cal.id, cal.puntaje, cal.comentario, cal.moderada, cal.fecha,
                  uf.nombre || ' ' || uf.apellido AS solicitante_nombre
           FROM calificaciones cal
           JOIN solicitantes f ON f.id = cal.solicitante_id
           JOIN usuarios uf    ON uf.id = f.usuario_id
           WHERE cal.prestador_id = ?
           ORDER BY cal.fecha DESC''',
        (pid,)
    ).fetchall()

    todos_verificados = (
        p['dni_verificado']  == 'VERIFICADO' and
        p['antecedentes_ok'] == 'VERIFICADO' and
        p['certificados_ok'] == 'VERIFICADO'
    )

    mes_inicio = datetime.now().strftime('%Y-%m-01')
    monto_mes_row = db.execute(
        """SELECT COALESCE(SUM(monto_neto), 0) AS total FROM pagos
           WHERE prestador_id=? AND estado='LIQUIDADO' AND fecha_pago >= ?""",
        (pid, mes_inicio)
    ).fetchone()
    monto_mes = round(monto_mes_row['total'], 2) if monto_mes_row else 0

    lim_row = db.execute(
        "SELECT valor FROM configuracion WHERE clave='factura_monto_limite'"
    ).fetchone()
    factura_limite = int(lim_row['valor']) if lim_row else 200000

    hoy_date  = date.today()
    hoy_iso   = hoy_date.isoformat()
    alert_iso = (hoy_date + timedelta(days=30)).isoformat()

    return render_template(
        'admin/prestadores/detalle.html',
        seccion_activa='prestadores',
        p=p,
        servicios=servicios,
        promedio=round(promedio, 1) if promedio else None,
        todos_verificados=todos_verificados,
        monto_mes=monto_mes,
        factura_limite=factura_limite,
        calificaciones_list=calificaciones_list,
        hoy_iso=hoy_iso,
        alert_iso=alert_iso,
        **_ctx()
    )


@admin_bp.route('/prestadores/<int:pid>/calificaciones/<int:rid>/moderar', methods=['POST'])
def prestador_calificacion_moderar(pid, rid):
    db = get_db()
    r  = db.execute('SELECT * FROM calificaciones WHERE id=? AND prestador_id=?', (rid, pid)).fetchone()
    if not r:
        flash('Calificación no encontrada.', 'error')
        return redirect(url_for('admin.prestador_detalle', pid=pid))
    nuevo = 0 if r['moderada'] else 1
    db.execute('UPDATE calificaciones SET moderada=? WHERE id=?', (nuevo, rid))
    db.commit()
    flash('Reseña moderada (oculta).' if nuevo else 'Reseña restaurada (visible).', 'success')
    return redirect(url_for('admin.prestador_detalle', pid=pid))


# ---------------------------------------------------------------------------
# Gestión de Prestadores — Nuevo
# ---------------------------------------------------------------------------

@admin_bp.route('/prestadores/nuevo', methods=['GET', 'POST'])
def prestador_nuevo():
    db         = get_db()
    categorias = db.execute('SELECT * FROM categorias WHERE activa=1 ORDER BY nombre').fetchall()
    zonas      = db.execute('SELECT * FROM zonas WHERE activa=1 ORDER BY nombre').fetchall()
    pwd        = generar_password_temporal()

    if request.method == 'POST':
        nombre      = request.form.get('nombre', '').strip()
        apellido    = request.form.get('apellido', '').strip()
        email       = request.form.get('email', '').strip().lower()
        telefono    = request.form.get('telefono', '').strip()
        categoria_id= request.form.get('categoria_id') or None
        zona_id     = request.form.get('zona_id') or None
        experiencia = int(request.form.get('experiencia_anios', '0') or 0)
        descripcion = request.form.get('descripcion', '').strip()
        password    = request.form.get('password_generada', '')

        errores = []
        if not nombre:   errores.append('El nombre es obligatorio.')
        if not apellido: errores.append('El apellido es obligatorio.')
        if not email:    errores.append('El email es obligatorio.')

        if not errores:
            if db.execute('SELECT id FROM usuarios WHERE email=?', (email,)).fetchone():
                errores.append('Ya existe un usuario con ese email.')

        if errores:
            for e in errores:
                flash(e, 'error')
            return render_template(
                'admin/prestadores/nuevo.html',
                seccion_activa='prestadores',
                categorias=categorias,
                zonas=zonas,
                password_generada=password,
                form=request.form,
                **_ctx()
            )

        hoy  = date.today().isoformat()
        venc = (date.today() + timedelta(days=90)).isoformat()

        usuario_id = db.execute(
            '''INSERT INTO usuarios
               (nombre, apellido, email, password_hash, tipo_usuario,
                estado, intentos_fallidos, fecha_alta, fecha_password,
                fecha_vencimiento, telefono)
               VALUES (?, ?, ?, ?, 'prestador', 'VENCIDA', 0, ?, ?, ?, ?)''',
            (nombre, apellido, email, generate_password_hash(password),
             datetime.now().isoformat(), hoy, venc, telefono or None)
        ).lastrowid

        prestador_id = db.execute(
            '''INSERT INTO prestadores
               (usuario_id, categoria_id, zona_id, experiencia_anios, descripcion,
                estado_perfil, dni_verificado, antecedentes_ok, certificados_ok)
               VALUES (?, ?, ?, ?, ?, 'EN_REVISION', 'PENDIENTE', 'PENDIENTE', 'PENDIENTE')''',
            (usuario_id, categoria_id, zona_id, experiencia, descripcion or None)
        ).lastrowid

        foto = request.files.get('foto')
        if foto and foto.filename:
            filename = _guardar_foto(foto, f'prestador_{prestador_id}')
            if filename:
                db.execute('UPDATE prestadores SET foto_url=? WHERE id=?',
                           (filename, prestador_id))

        db.commit()
        flash(
            f'Prestador {nombre} {apellido} creado correctamente. '
            f'Contraseña temporal: {password}',
            'success'
        )
        return redirect(url_for('admin.prestadores'))

    return render_template(
        'admin/prestadores/nuevo.html',
        seccion_activa='prestadores',
        categorias=categorias,
        zonas=zonas,
        password_generada=pwd,
        form={},
        **_ctx()
    )


# ---------------------------------------------------------------------------
# Gestión de Prestadores — Editar
# ---------------------------------------------------------------------------

@admin_bp.route('/prestadores/<int:pid>/editar', methods=['GET', 'POST'])
def prestador_editar(pid):
    db = get_db()
    p  = _get_prestador(db, pid)
    if not p:
        flash('Prestador no encontrado.', 'error')
        return redirect(url_for('admin.prestadores'))

    categorias = db.execute('SELECT * FROM categorias WHERE activa=1 ORDER BY nombre').fetchall()
    zonas      = db.execute('SELECT * FROM zonas WHERE activa=1 ORDER BY nombre').fetchall()

    if request.method == 'POST':
        nombre        = request.form.get('nombre', '').strip()
        apellido      = request.form.get('apellido', '').strip()
        telefono      = request.form.get('telefono', '').strip()
        categoria_id  = request.form.get('categoria_id') or None
        zona_id       = request.form.get('zona_id') or None
        experiencia   = int(request.form.get('experiencia_anios', '0') or 0)
        descripcion   = request.form.get('descripcion', '').strip()
        codigo_postal = request.form.get('codigo_postal', '').strip() or None
        localidad     = request.form.get('localidad', '').strip() or None
        provincia     = request.form.get('provincia', '').strip() or None
        try:
            tarifa_hora = float(request.form.get('tarifa_hora', '0').replace(',', '.'))
        except (ValueError, AttributeError):
            tarifa_hora = 0.0

        if not nombre or not apellido:
            flash('Nombre y apellido son obligatorios.', 'error')
            return render_template(
                'admin/prestadores/editar.html',
                seccion_activa='prestadores',
                p=p, categorias=categorias, zonas=zonas,
                **_ctx()
            )

        db.execute(
            'UPDATE usuarios SET nombre=?, apellido=?, telefono=? WHERE id=?',
            (nombre, apellido, telefono or None, p['usuario_id'])
        )
        db.execute(
            '''UPDATE prestadores
               SET categoria_id=?, zona_id=?, experiencia_anios=?, descripcion=?,
                   codigo_postal=?, localidad=?, provincia=?, tarifa_hora=?
               WHERE id=?''',
            (categoria_id, zona_id, experiencia, descripcion or None, codigo_postal, localidad, provincia, tarifa_hora, pid)
        )

        foto = request.files.get('foto')
        if foto and foto.filename:
            filename = _guardar_foto(foto, f'prestador_{pid}')
            if filename:
                db.execute('UPDATE prestadores SET foto_url=? WHERE id=?', (filename, pid))

        db.commit()
        flash('Prestador actualizado correctamente.', 'success')
        return redirect(url_for('admin.prestador_detalle', pid=pid))

    return render_template(
        'admin/prestadores/editar.html',
        seccion_activa='prestadores',
        p=p, categorias=categorias, zonas=zonas,
        **_ctx()
    )


# ---------------------------------------------------------------------------
# Verificación de documentación
# ---------------------------------------------------------------------------

DOC_CAMPOS = {
    'dni':          ('dni_verificado',  'tu DNI'),
    'antecedentes': ('antecedentes_ok', 'tu certificado de antecedentes'),
    'certificados': ('certificados_ok', 'tus certificados de formación'),
}


@admin_bp.route('/prestadores/<int:pid>/doc/<doc>/aprobar', methods=['POST'])
def prestador_doc_aprobar(pid, doc):
    if doc not in DOC_CAMPOS:
        flash('Documento no válido.', 'error')
        return redirect(url_for('admin.prestador_detalle', pid=pid))

    campo, nombre_doc = DOC_CAMPOS[doc]
    db = get_db()
    p  = db.execute('SELECT * FROM prestadores WHERE id=?', (pid,)).fetchone()
    if not p:
        flash('Prestador no encontrado.', 'error')
        return redirect(url_for('admin.prestadores'))

    if doc == 'dni':
        if not p['numero_dni']:
            flash('No podés aprobar el DNI: el prestador no ingresó su número.', 'error')
            return redirect(url_for('admin.prestador_detalle', pid=pid))
        if not p['dni_foto_frente_url'] or not p['dni_foto_selfie_url']:
            flash('No podés aprobar el DNI: faltan las fotos del documento (frente y/o selfie).', 'error')
            return redirect(url_for('admin.prestador_detalle', pid=pid))

    db.execute(f"UPDATE prestadores SET {campo}='VERIFICADO' WHERE id=?", (pid,))
    _notificar(db, p['usuario_id'], 'DOC',
               f'Tu {nombre_doc} fue verificado/a correctamente')

    # Verificar si los 3 documentos quedaron verificados
    p2       = db.execute('SELECT * FROM prestadores WHERE id=?', (pid,)).fetchone()
    todos_ok = (p2['dni_verificado'] == 'VERIFICADO' and
                p2['antecedentes_ok'] == 'VERIFICADO' and
                p2['certificados_ok'] == 'VERIFICADO')
    db.commit()

    msg = 'Documento aprobado correctamente.'
    if todos_ok and p2['estado_perfil'] != 'APROBADO':
        msg += ' ¡Los 3 documentos están verificados! Podés aprobar el perfil completo.'
    flash(msg, 'success')
    return redirect(url_for('admin.prestador_detalle', pid=pid))


@admin_bp.route('/prestadores/<int:pid>/doc/<doc>/rechazar', methods=['POST'])
def prestador_doc_rechazar(pid, doc):
    if doc not in DOC_CAMPOS:
        flash('Documento no válido.', 'error')
        return redirect(url_for('admin.prestador_detalle', pid=pid))

    campo, nombre_doc = DOC_CAMPOS[doc]
    motivo = request.form.get('motivo', '').strip()
    if not motivo:
        flash('El motivo de rechazo es obligatorio.', 'error')
        return redirect(url_for('admin.prestador_detalle', pid=pid))

    db = get_db()
    p  = db.execute('SELECT * FROM prestadores WHERE id=?', (pid,)).fetchone()
    if not p:
        flash('Prestador no encontrado.', 'error')
        return redirect(url_for('admin.prestadores'))

    db.execute(f"UPDATE prestadores SET {campo}='RECHAZADO' WHERE id=?", (pid,))
    _notificar(db, p['usuario_id'], 'DOC',
               f'Tu {nombre_doc} fue rechazado/a. Motivo: {motivo}')
    db.commit()
    flash('Documento rechazado.', 'success')
    return redirect(url_for('admin.prestador_detalle', pid=pid))


# ---------------------------------------------------------------------------
# Antecedentes — subida de PDF y alertas de vencimiento
# ---------------------------------------------------------------------------

def _enviar_alerta_antecedentes(db, prestador_id, fecha_venc):
    """Notifica al prestador y al admin que el certificado está por vencer."""
    pr = db.execute(
        '''SELECT p.usuario_id, u.nombre, u.apellido
           FROM prestadores p
           JOIN usuarios u ON p.usuario_id = u.id
           WHERE p.id = ?''',
        (prestador_id,)
    ).fetchone()
    if not pr:
        return

    dias_restantes = (fecha_venc - date.today()).days
    fecha_str = fecha_venc.strftime('%d/%m/%Y')

    _notificar(db, pr['usuario_id'], 'antecedentes_por_vencer',
               'Tu certificado de antecedentes vence pronto',
               f'Tu certificado de antecedentes penales vence el {fecha_str} '
               f'(en {dias_restantes} días). Solicitá la renovación en '
               f'argentina.gob.ar/justicia/reincidencia/antecedentespenales '
               f'y envialo a antecedentes@amparo.com')

    admins = db.execute(
        "SELECT id FROM usuarios WHERE tipo_usuario IN ('admin', 'admin_financiero')"
    ).fetchall()
    for admin in admins:
        _notificar(db, admin['id'], 'antecedentes_por_vencer',
                   f"Certificado por vencer — {pr['nombre']} {pr['apellido']}",
                   f"El certificado de antecedentes de {pr['nombre']} {pr['apellido']} "
                   f"vence el {fecha_str}. Recordale que lo renueve.")

    db.execute(
        'UPDATE prestadores SET antecedentes_alerta_enviada = 1 WHERE id = ?',
        (prestador_id,)
    )


@admin_bp.route('/prestadores/<int:pid>/antecedentes/subir', methods=['POST'])
def prestador_antecedentes_subir(pid):
    db = get_db()
    p  = _get_prestador(db, pid)
    if not p:
        flash('Prestador no encontrado.', 'error')
        return redirect(url_for('admin.prestadores'))

    archivo       = request.files.get('antecedentes_pdf')
    fecha_emision = request.form.get('fecha_emision', '').strip()

    if not archivo or not archivo.filename:
        flash('Debés seleccionar un archivo PDF.', 'error')
        return redirect(url_for('admin.prestador_detalle', pid=pid))
    if not archivo.filename.lower().endswith('.pdf'):
        flash('El archivo debe ser un PDF.', 'error')
        return redirect(url_for('admin.prestador_detalle', pid=pid))
    if not fecha_emision:
        flash('La fecha de emisión es obligatoria.', 'error')
        return redirect(url_for('admin.prestador_detalle', pid=pid))

    try:
        fe = date.fromisoformat(fecha_emision)
    except ValueError:
        flash('Fecha de emisión inválida.', 'error')
        return redirect(url_for('admin.prestador_detalle', pid=pid))

    # Calcular vencimiento
    _row_meses = db.execute(
        "SELECT valor FROM configuracion WHERE clave='antecedentes_vigencia_meses'"
    ).fetchone()
    meses = int(_row_meses['valor'] if _row_meses and _row_meses['valor'] else 12)
    mes_venc = fe.month - 1 + meses
    anio_venc = fe.year + mes_venc // 12
    mes_venc = mes_venc % 12 + 1
    import calendar
    ultimo_dia = calendar.monthrange(anio_venc, mes_venc)[1]
    fv = date(anio_venc, mes_venc, min(fe.day, ultimo_dia))

    # Guardar archivo
    import time
    ts       = int(time.time())
    filename = f'ant_{pid}_{ts}.pdf'
    os.makedirs(ANTECEDENTES_FOLDER, exist_ok=True)
    archivo.save(os.path.join(ANTECEDENTES_FOLDER, filename))
    pdf_url  = filename

    db.execute(
        '''UPDATE prestadores SET
               antecedentes_ok = 'VERIFICADO',
               antecedentes_pdf_url = ?,
               antecedentes_fecha_emision = ?,
               antecedentes_fecha_vencimiento = ?,
               antecedentes_alerta_enviada = 0
           WHERE id = ?''',
        (pdf_url, fe.isoformat(), fv.isoformat(), pid)
    )

    _notificar(db, p['usuario_id'], 'DOC',
               'Tu certificado de antecedentes fue verificado y archivado',
               f'Tu certificado de antecedentes penales fue verificado y archivado. '
               f'Vence el {fv.strftime("%d/%m/%Y")}.')
    db.commit()
    flash(f'Certificado subido correctamente. Vence el {fv.strftime("%d/%m/%Y")}.', 'success')
    return redirect(url_for('admin.prestador_detalle', pid=pid))


@admin_bp.route('/prestadores/<int:pid>/antecedentes/ver')
def prestador_antecedentes_ver(pid):
    from flask import send_file
    db = get_db()
    p  = db.execute('SELECT antecedentes_pdf_url FROM prestadores WHERE id=?', (pid,)).fetchone()
    if not p or not p['antecedentes_pdf_url']:
        flash('No hay PDF de antecedentes para este prestador.', 'error')
        return redirect(url_for('admin.prestador_detalle', pid=pid))
    ruta = os.path.join(ANTECEDENTES_FOLDER, p['antecedentes_pdf_url'])
    return send_file(ruta, mimetype='application/pdf')


# ---------------------------------------------------------------------------
# CV y factura del prestador
# ---------------------------------------------------------------------------

@admin_bp.route('/prestadores/<int:pid>/cv-recibido', methods=['POST'])
def prestador_cv_recibido(pid):
    db = get_db()
    p  = db.execute('SELECT id FROM prestadores WHERE id=?', (pid,)).fetchone()
    if not p:
        flash('Prestador no encontrado.', 'error')
        return redirect(url_for('admin.prestadores'))
    db.execute('UPDATE prestadores SET cv_email_enviado=1 WHERE id=?', (pid,))
    db.commit()
    flash('CV marcado como recibido por email.', 'success')
    return redirect(url_for('admin.prestador_detalle', pid=pid))


@admin_bp.route('/prestadores/<int:pid>/factura-recibida', methods=['POST'])
def prestador_factura_recibida(pid):
    db = get_db()
    p  = db.execute('SELECT * FROM prestadores WHERE id=?', (pid,)).fetchone()
    if not p:
        flash('Prestador no encontrado.', 'error')
        return redirect(url_for('admin.prestadores'))
    db.execute('UPDATE prestadores SET factura_requerida=0 WHERE id=?', (pid,))
    _notificar(db, p['usuario_id'], 'FISCAL',
               'El equipo de AMPARO confirmó la recepción de tu factura.')
    db.commit()
    flash('Factura marcada como recibida.', 'success')
    return redirect(url_for('admin.prestador_detalle', pid=pid))


# ---------------------------------------------------------------------------
# Estado del perfil
# ---------------------------------------------------------------------------

@admin_bp.route('/prestadores/<int:pid>/aprobar', methods=['POST'])
def prestador_aprobar(pid):
    db = get_db()
    p  = db.execute('SELECT * FROM prestadores WHERE id=?', (pid,)).fetchone()
    if not p:
        flash('Prestador no encontrado.', 'error')
        return redirect(url_for('admin.prestadores'))
    if not p['numero_dni'] or not p['dni_foto_frente_url'] or not p['dni_foto_selfie_url']:
        flash('No podés aprobar el perfil: faltan el número de DNI o las fotos del documento.', 'error')
        return redirect(url_for('admin.prestador_detalle', pid=pid))
    if p['antecedentes_ok'] == 'PENDIENTE':
        flash('No podés aprobar el perfil sin verificar primero los antecedentes penales.', 'error')
        return redirect(url_for('admin.prestador_detalle', pid=pid))
    db.execute(
        "UPDATE prestadores SET estado_perfil='APROBADO', fecha_aprobacion=? WHERE id=?",
        (datetime.now().isoformat(), pid)
    )
    _notificar(db, p['usuario_id'], 'PERFIL',
               'Tu perfil fue aprobado. Ya sos visible para los solicitantes.')
    db.commit()
    u = db.execute('SELECT nombre, email FROM usuarios WHERE id=?', (p['usuario_id'],)).fetchone()
    if u:
        app_url   = _cfg_db('app_url', 'http://127.0.0.1:5000')
        link_app  = app_url + '/prestador/dashboard'
        asunto    = _cfg_db('mail_perfil_aprobado_asunto', '¡Tu perfil fue aprobado! — AMPARO Red')
        cuerpo    = _cfg_db('mail_perfil_aprobado_cuerpo',
            'Hola {nombre},\n\nTu perfil fue aprobado. '
            'Ya sos visible para los solicitantes.\n\n{link_app}')
        cuerpo    = cuerpo.replace('{nombre}', u['nombre']).replace('{link_app}', link_app)
        import os as _os
        base_dir    = _os.path.dirname(_os.path.abspath(__file__))
        ruta_manual = _os.path.join(base_dir, '..', 'static', 'docs', 'manual_prestador.pdf')
        enviar_email(u['email'], asunto, cuerpo,
                     adjunto_path=ruta_manual,
                     adjunto_nombre='Manual_AMPARO_Red.pdf')
    flash('Perfil aprobado correctamente.', 'success')
    return redirect(url_for('admin.prestador_detalle', pid=pid))


@admin_bp.route('/prestadores/<int:pid>/rechazar', methods=['POST'])
def prestador_rechazar(pid):
    motivo = request.form.get('motivo', '').strip()
    if not motivo:
        flash('El motivo de rechazo es obligatorio.', 'error')
        return redirect(url_for('admin.prestador_detalle', pid=pid))
    print(f'[RECHAZO] Procesando rechazo para prestador {pid}')
    db = get_db()
    p  = db.execute('SELECT * FROM prestadores WHERE id=?', (pid,)).fetchone()
    if not p:
        flash('Prestador no encontrado.', 'error')
        return redirect(url_for('admin.prestadores'))
    db.execute(
        "UPDATE prestadores SET estado_perfil='RECHAZADO', motivo_rechazo=? WHERE id=?",
        (motivo, pid)
    )
    _notificar(db, p['usuario_id'], 'PERFIL',
               f'Tu perfil fue rechazado. Motivo: {motivo}')
    db.commit()
    u = db.execute('SELECT nombre, email FROM usuarios WHERE id=?', (p['usuario_id'],)).fetchone()
    if u:
        print(f'[RECHAZO] Enviando correo a {u["email"]}')
        asunto = _cfg_db('mail_perfil_rechazado_asunto', 'Actualización sobre tu perfil — AMPARO Red')
        cuerpo = _cfg_db('mail_perfil_rechazado_cuerpo',
            'Hola {nombre},\n\nTu perfil fue rechazado.\n\nMotivo: {motivo_rechazo}\n\n'
            'Podés corregir la información y volver a enviar tu solicitud.')
        cuerpo = cuerpo.replace('{nombre}', u['nombre']).replace('{motivo_rechazo}', motivo)
        enviar_email(u['email'], asunto, cuerpo)
    flash('Perfil rechazado.', 'success')
    return redirect(url_for('admin.prestador_detalle', pid=pid))


@admin_bp.route('/prestadores/<int:pid>/suspender', methods=['POST'])
def prestador_suspender(pid):
    motivo = request.form.get('motivo', '').strip()
    if not motivo:
        flash('El motivo de suspensión es obligatorio.', 'error')
        return redirect(url_for('admin.prestador_detalle', pid=pid))
    db = get_db()
    p  = db.execute('SELECT * FROM prestadores WHERE id=?', (pid,)).fetchone()
    if not p:
        flash('Prestador no encontrado.', 'error')
        return redirect(url_for('admin.prestadores'))
    db.execute("UPDATE prestadores SET estado_perfil='SUSPENDIDO' WHERE id=?", (pid,))
    _notificar(db, p['usuario_id'], 'PERFIL',
               f'Tu perfil fue suspendido. Motivo: {motivo}')
    db.commit()
    flash('Prestador suspendido correctamente.', 'success')
    return redirect(url_for('admin.prestador_detalle', pid=pid))


@admin_bp.route('/prestadores/<int:pid>/reactivar', methods=['POST'])
def prestador_reactivar(pid):
    db = get_db()
    p  = db.execute('SELECT * FROM prestadores WHERE id=?', (pid,)).fetchone()
    if not p:
        flash('Prestador no encontrado.', 'error')
        return redirect(url_for('admin.prestadores'))
    db.execute("UPDATE prestadores SET estado_perfil='APROBADO' WHERE id=?", (pid,))
    _notificar(db, p['usuario_id'], 'PERFIL',
               'Tu perfil fue reactivado. Ya sos visible nuevamente para los solicitantes.')
    db.commit()
    flash('Prestador reactivado correctamente.', 'success')
    return redirect(url_for('admin.prestador_detalle', pid=pid))


# ---------------------------------------------------------------------------
# Secciones del menú (en construcción)
# ---------------------------------------------------------------------------

def _construccion(seccion, titulo):
    return render_template(
        'admin/construccion.html',
        seccion_activa=seccion,
        titulo=titulo,
        **_ctx()
    )


# ---------------------------------------------------------------------------
# Gestión de Familias — helpers
# ---------------------------------------------------------------------------

def _get_solicitante(db, fid):
    return db.execute(
        '''SELECT f.id, f.usuario_id, f.zona_id, f.direccion,
                  f.latitud, f.longitud, f.codigo_postal, f.localidad, f.provincia,
                  f.metodo_pago, f.metodo_pago_descripcion, f.mp_card_token,
                  f.familiar_nombre, f.familiar_edad,
                  f.familiar_condicion, f.familiar_necesidades,
                  u.nombre, u.apellido, u.email, u.telefono,
                  u.estado, u.fecha_alta, u.ultimo_ingreso,
                  z.nombre AS zona_nombre
           FROM solicitantes f
           JOIN usuarios u ON f.usuario_id = u.id
           LEFT JOIN zonas z ON f.zona_id = z.id
           WHERE f.id = ?''',
        (fid,)
    ).fetchone()


# ---------------------------------------------------------------------------
# Gestión de Familias — Listado
# ---------------------------------------------------------------------------

@admin_bp.route('/solicitantes')
def solicitantes():
    busqueda = request.args.get('busqueda', '').strip()
    zona_id  = request.args.get('zona_id', '')
    estado   = request.args.get('estado', '')

    db     = get_db()
    query  = '''
        SELECT f.*, u.nombre, u.apellido, u.email, u.telefono,
               u.estado, u.fecha_alta, u.ultimo_ingreso,
               z.nombre AS zona_nombre,
               COUNT(CASE WHEN s.estado = 'FINALIZADO' THEN 1 END) AS total_contrataciones
        FROM solicitantes f
        JOIN usuarios u ON f.usuario_id = u.id
        LEFT JOIN zonas z ON f.zona_id = z.id
        LEFT JOIN servicios s ON s.solicitante_id = f.id
        WHERE 1=1
    '''
    params = []

    if busqueda:
        query += ' AND (u.nombre LIKE ? OR u.apellido LIKE ? OR u.email LIKE ?)'
        params += [f'%{busqueda}%', f'%{busqueda}%', f'%{busqueda}%']
    if zona_id:
        query += ' AND f.zona_id = ?'
        params.append(int(zona_id))
    if estado:
        query += ' AND u.estado = ?'
        params.append(estado)

    query += ' GROUP BY f.id ORDER BY u.fecha_alta DESC'

    lista = db.execute(query, params).fetchall()
    zonas = db.execute('SELECT * FROM zonas WHERE activa=1 ORDER BY nombre').fetchall()

    return render_template(
        'admin/solicitantes/listado.html',
        seccion_activa='solicitantes',
        solicitantes=lista,
        zonas=zonas,
        busqueda=busqueda,
        zona_id=zona_id,
        estado=estado,
        **_ctx()
    )


# ---------------------------------------------------------------------------
# Gestión de Familias — Detalle
# ---------------------------------------------------------------------------

@admin_bp.route('/solicitantes/<int:fid>')
def solicitante_detalle(fid):
    db = get_db()
    f  = _get_solicitante(db, fid)
    if not f:
        flash('Familia no encontrada.', 'error')
        return redirect(url_for('admin.solicitantes'))

    servicios = db.execute(
        '''SELECT s.fecha_servicio, s.monto_acordado, s.estado, s.id AS servicio_id,
                  u.nombre || ' ' || u.apellido AS prestador_nombre,
                  c.nombre AS categoria_nombre,
                  cal.puntaje
           FROM servicios s
           JOIN prestadores p   ON s.prestador_id = p.id
           JOIN usuarios u      ON p.usuario_id   = u.id
           LEFT JOIN categorias c   ON p.categoria_id = c.id
           LEFT JOIN calificaciones cal ON s.id = cal.servicio_id
           WHERE s.solicitante_id = ?
           ORDER BY s.fecha_servicio DESC LIMIT 20''',
        (fid,)
    ).fetchall()

    resenas = db.execute(
        '''SELECT cal.id, cal.puntaje, cal.comentario, cal.moderada, cal.fecha,
                  u.nombre || ' ' || u.apellido AS prestador_nombre,
                  s.estado AS estado_servicio,
                  s.fecha_servicio
           FROM calificaciones cal
           JOIN prestadores p  ON cal.prestador_id = p.id
           JOIN usuarios u     ON p.usuario_id     = u.id
           JOIN servicios s    ON s.id             = cal.servicio_id
           WHERE cal.solicitante_id = ?
           ORDER BY cal.fecha DESC''',
        (fid,)
    ).fetchall()

    pagos_tabla = db.execute(
        '''SELECT DISTINCT metodo_pago, fecha_pago
           FROM pagos
           WHERE solicitante_id = ? AND metodo_pago IS NOT NULL
           ORDER BY fecha_pago DESC''',
        (fid,)
    ).fetchall()

    reclamos = db.execute(
        '''SELECT r.id, r.descripcion, r.estado, r.resolucion,
                  r.fecha_apertura, r.fecha_resolucion
           FROM reclamos r
           JOIN servicios s ON r.servicio_id = s.id
           WHERE s.solicitante_id = ?
           ORDER BY r.fecha_apertura DESC''',
        (fid,)
    ).fetchall()

    return render_template(
        'admin/solicitantes/detalle.html',
        seccion_activa='solicitantes',
        f=f,
        servicios=servicios,
        resenas=resenas,
        pagos_tabla=pagos_tabla,
        reclamos=reclamos,
        **_ctx()
    )


# ---------------------------------------------------------------------------
# Gestión de Familias — Nueva
# ---------------------------------------------------------------------------

@admin_bp.route('/solicitantes/nueva', methods=['GET', 'POST'])
def solicitante_nueva():
    db    = get_db()
    zonas = db.execute('SELECT * FROM zonas WHERE activa=1 ORDER BY nombre').fetchall()
    pwd   = generar_password_temporal()

    if request.method == 'POST':
        nombre      = request.form.get('nombre', '').strip()
        apellido    = request.form.get('apellido', '').strip()
        email       = request.form.get('email', '').strip().lower()
        telefono    = request.form.get('telefono', '').strip()
        direccion   = request.form.get('direccion', '').strip()
        zona_id     = request.form.get('zona_id') or None
        fam_nombre  = request.form.get('familiar_nombre', '').strip()
        fam_edad    = request.form.get('familiar_edad', '').strip()
        fam_condic  = request.form.get('familiar_condicion', '').strip()
        fam_nec     = request.form.get('familiar_necesidades', '').strip()
        password    = request.form.get('password_generada', '')

        errores = []
        if not nombre:   errores.append('El nombre es obligatorio.')
        if not apellido: errores.append('El apellido es obligatorio.')
        if not email:    errores.append('El email es obligatorio.')

        if not errores:
            if db.execute('SELECT id FROM usuarios WHERE email=?', (email,)).fetchone():
                errores.append('Ya existe un usuario con ese email.')

        if errores:
            for e in errores:
                flash(e, 'error')
            return render_template(
                'admin/solicitantes/nuevo.html',
                seccion_activa='solicitantes',
                zonas=zonas, password_generada=password, form=request.form,
                **_ctx()
            )

        hoy  = date.today().isoformat()
        venc = (date.today() + timedelta(days=90)).isoformat()

        uid = db.execute(
            '''INSERT INTO usuarios
               (nombre, apellido, email, password_hash, tipo_usuario,
                estado, intentos_fallidos, fecha_alta, fecha_password,
                fecha_vencimiento, telefono)
               VALUES (?, ?, ?, ?, 'solicitante', 'VENCIDA', 0, ?, ?, ?, ?)''',
            (nombre, apellido, email, generate_password_hash(password),
             datetime.now().isoformat(), hoy, venc, telefono or None)
        ).lastrowid

        db.execute(
            '''INSERT INTO solicitantes
               (usuario_id, direccion, zona_id, familiar_nombre,
                familiar_edad, familiar_condicion, familiar_necesidades)
               VALUES (?, ?, ?, ?, ?, ?, ?)''',
            (uid, direccion or None, zona_id,
             fam_nombre or None, int(fam_edad) if fam_edad.isdigit() else None,
             fam_condic or None, fam_nec or None)
        )
        db.commit()
        flash(
            f'Familia {nombre} {apellido} creada correctamente. '
            f'Contraseña temporal: {password}',
            'success'
        )
        return redirect(url_for('admin.solicitantes'))

    return render_template(
        'admin/solicitantes/nuevo.html',
        seccion_activa='solicitantes',
        zonas=zonas, password_generada=pwd, form={},
        **_ctx()
    )


# ---------------------------------------------------------------------------
# Gestión de Familias — Editar
# ---------------------------------------------------------------------------

@admin_bp.route('/solicitantes/<int:fid>/editar', methods=['GET', 'POST'])
def solicitante_editar(fid):
    db = get_db()
    f  = _get_solicitante(db, fid)
    if not f:
        flash('Familia no encontrada.', 'error')
        return redirect(url_for('admin.solicitantes'))

    zonas = db.execute('SELECT * FROM zonas WHERE activa=1 ORDER BY nombre').fetchall()

    if request.method == 'POST':
        nombre     = request.form.get('nombre', '').strip()
        apellido   = request.form.get('apellido', '').strip()
        telefono   = request.form.get('telefono', '').strip()
        direccion  = request.form.get('direccion', '').strip()
        zona_id    = request.form.get('zona_id') or None
        fam_nombre = request.form.get('familiar_nombre', '').strip()
        fam_edad   = request.form.get('familiar_edad', '').strip()
        fam_condic = request.form.get('familiar_condicion', '').strip()
        fam_nec    = request.form.get('familiar_necesidades', '').strip()

        if not nombre or not apellido:
            flash('Nombre y apellido son obligatorios.', 'error')
            return render_template(
                'admin/solicitantes/editar.html',
                seccion_activa='solicitantes', f=f, zonas=zonas, **_ctx()
            )

        db.execute(
            'UPDATE usuarios SET nombre=?, apellido=?, telefono=? WHERE id=?',
            (nombre, apellido, telefono or None, f['usuario_id'])
        )
        db.execute(
            '''UPDATE solicitantes
               SET direccion=?, zona_id=?, familiar_nombre=?,
                   familiar_edad=?, familiar_condicion=?, familiar_necesidades=?
               WHERE id=?''',
            (direccion or None, zona_id,
             fam_nombre or None, int(fam_edad) if fam_edad.isdigit() else None,
             fam_condic or None, fam_nec or None, fid)
        )
        db.commit()
        flash('Familia actualizada correctamente.', 'success')
        return redirect(url_for('admin.solicitante_detalle', fid=fid))

    return render_template(
        'admin/solicitantes/editar.html',
        seccion_activa='solicitantes', f=f, zonas=zonas, **_ctx()
    )


# ---------------------------------------------------------------------------
# Acciones directas sobre solicitantes
# ---------------------------------------------------------------------------

@admin_bp.route('/solicitantes/<int:fid>/desactivar', methods=['POST'])
def solicitante_desactivar(fid):
    db = get_db()
    f  = _get_solicitante(db, fid)
    if not f:
        flash('Familia no encontrada.', 'error')
        return redirect(url_for('admin.solicitantes'))
    db.execute("UPDATE usuarios SET estado='INACTIVA' WHERE id=?", (f['usuario_id'],))
    db.commit()
    flash('Cuenta desactivada correctamente.', 'success')
    return redirect(url_for('admin.solicitante_detalle', fid=fid))


@admin_bp.route('/solicitantes/<int:fid>/reactivar', methods=['POST'])
def solicitante_reactivar(fid):
    db = get_db()
    f  = _get_solicitante(db, fid)
    if not f:
        flash('Familia no encontrada.', 'error')
        return redirect(url_for('admin.solicitantes'))
    db.execute("UPDATE usuarios SET estado='ACTIVA' WHERE id=?", (f['usuario_id'],))
    db.commit()
    flash('Cuenta reactivada correctamente.', 'success')
    return redirect(url_for('admin.solicitante_detalle', fid=fid))


@admin_bp.route('/solicitantes/<int:fid>/resenas/<int:rid>/moderar', methods=['POST'])
def solicitante_resena_moderar(fid, rid):
    db = get_db()
    r  = db.execute('SELECT * FROM calificaciones WHERE id=?', (rid,)).fetchone()
    if not r:
        flash('Reseña no encontrada.', 'error')
        return redirect(url_for('admin.solicitante_detalle', fid=fid))
    nuevo = 0 if r['moderada'] else 1
    db.execute('UPDATE calificaciones SET moderada=? WHERE id=?', (nuevo, rid))
    db.commit()
    msg = 'Reseña moderada (oculta).' if nuevo else 'Reseña restaurada (visible).'
    flash(msg, 'success')
    return redirect(url_for('admin.solicitante_detalle', fid=fid))


@admin_bp.route('/solicitantes/<int:fid>/reclamos/<int:rid>/gestionar', methods=['POST'])
def solicitante_reclamo_gestionar(fid, rid):
    estado     = request.form.get('estado', '').strip()
    resolucion = request.form.get('resolucion', '').strip()
    if not estado:
        flash('El estado es obligatorio.', 'error')
        return redirect(url_for('admin.solicitante_detalle', fid=fid))
    db = get_db()
    fecha_res = datetime.now().isoformat() if estado == 'RESUELTO' else None
    db.execute(
        'UPDATE reclamos SET estado=?, resolucion=?, fecha_resolucion=? WHERE id=?',
        (estado, resolucion or None, fecha_res, rid)
    )
    db.commit()
    flash('Reclamo actualizado correctamente.', 'success')
    return redirect(url_for('admin.solicitante_detalle', fid=fid))


# ---------------------------------------------------------------------------
# Gestión de Categorías
# ---------------------------------------------------------------------------

CATEGORIAS_PROTEGIDAS = {'Cuidador Domiciliario', 'Acompañante Terapéutico', 'Enfermero Domiciliario'}


@admin_bp.route('/categorias')
def categorias():
    estado = request.args.get('estado', '')
    db = get_db()

    query = '''
        SELECT c.*,
               COUNT(p.id) AS prestadores_activos
        FROM categorias c
        LEFT JOIN prestadores p ON p.categoria_id = c.id AND p.estado_perfil = 'APROBADO'
        WHERE 1=1
    '''
    params = []
    if estado == 'activas':
        query += ' AND c.activa = 1'
    elif estado == 'inactivas':
        query += ' AND c.activa = 0'
    query += ' GROUP BY c.id ORDER BY c.nombre'

    lista = db.execute(query, params).fetchall()

    return render_template(
        'admin/categorias/listado.html',
        seccion_activa='categorias',
        categorias=lista,
        estado=estado,
        **_ctx()
    )


@admin_bp.route('/categorias/nueva', methods=['GET', 'POST'])
def categoria_nueva():
    if request.method == 'POST':
        nombre     = request.form.get('nombre', '').strip()
        desc       = request.form.get('descripcion', '').strip()
        requisitos = request.form.get('requisitos', '').strip()
        t_min      = request.form.get('tarifa_minima', '').strip()
        t_max      = request.form.get('tarifa_maxima', '').strip()

        errores = []
        if not nombre: errores.append('El nombre es obligatorio.')
        if not desc:   errores.append('La descripción es obligatoria.')
        if not t_min:  errores.append('La tarifa mínima es obligatoria.')
        if not t_max:  errores.append('La tarifa máxima es obligatoria.')

        t_min_f = t_max_f = None
        if not errores:
            try:
                t_min_f = float(t_min)
                t_max_f = float(t_max)
            except ValueError:
                errores.append('Las tarifas deben ser valores numéricos.')

        if not errores and t_max_f <= t_min_f:
            errores.append('La tarifa máxima debe ser mayor que la mínima.')

        if not errores:
            db = get_db()
            if db.execute('SELECT id FROM categorias WHERE nombre = ?', (nombre,)).fetchone():
                errores.append('Ya existe una categoría con ese nombre.')

        if errores:
            for e in errores:
                flash(e, 'error')
            return render_template(
                'admin/categorias/nueva.html',
                seccion_activa='categorias',
                form=request.form,
                **_ctx()
            )

        db = get_db()
        db.execute(
            'INSERT INTO categorias (nombre, descripcion, requisitos, tarifa_minima, tarifa_maxima, activa) VALUES (?, ?, ?, ?, ?, 1)',
            (nombre, desc, requisitos or None, t_min_f, t_max_f)
        )
        db.commit()
        flash(f'Categoría "{nombre}" creada correctamente.', 'success')
        return redirect(url_for('admin.categorias'))

    return render_template(
        'admin/categorias/nueva.html',
        seccion_activa='categorias',
        form={},
        **_ctx()
    )


@admin_bp.route('/categorias/<int:cid>')
def categoria_detalle(cid):
    db  = get_db()
    cat = db.execute('SELECT * FROM categorias WHERE id = ?', (cid,)).fetchone()
    if not cat:
        flash('Categoría no encontrada.', 'error')
        return redirect(url_for('admin.categorias'))

    prestadores_cat = db.execute(
        '''SELECT p.id, u.nombre, u.apellido, p.estado_perfil,
                  z.nombre AS zona_nombre
           FROM prestadores p
           JOIN usuarios u ON p.usuario_id = u.id
           LEFT JOIN zonas z ON p.zona_id = z.id
           WHERE p.categoria_id = ? AND p.estado_perfil = 'APROBADO'
           ORDER BY u.apellido, u.nombre''',
        (cid,)
    ).fetchall()

    return render_template(
        'admin/categorias/detalle.html',
        seccion_activa='categorias',
        cat=cat,
        prestadores=prestadores_cat,
        es_protegida=cat['nombre'] in CATEGORIAS_PROTEGIDAS,
        **_ctx()
    )


@admin_bp.route('/categorias/<int:cid>/editar', methods=['GET', 'POST'])
def categoria_editar(cid):
    db  = get_db()
    cat = db.execute('SELECT * FROM categorias WHERE id = ?', (cid,)).fetchone()
    if not cat:
        flash('Categoría no encontrada.', 'error')
        return redirect(url_for('admin.categorias'))

    if request.method == 'POST':
        nombre     = request.form.get('nombre', '').strip()
        desc       = request.form.get('descripcion', '').strip()
        requisitos = request.form.get('requisitos', '').strip()
        t_min      = request.form.get('tarifa_minima', '').strip()
        t_max      = request.form.get('tarifa_maxima', '').strip()

        errores = []
        if not nombre: errores.append('El nombre es obligatorio.')
        if not desc:   errores.append('La descripción es obligatoria.')
        if not t_min:  errores.append('La tarifa mínima es obligatoria.')
        if not t_max:  errores.append('La tarifa máxima es obligatoria.')

        t_min_f = t_max_f = None
        if not errores:
            try:
                t_min_f = float(t_min)
                t_max_f = float(t_max)
            except ValueError:
                errores.append('Las tarifas deben ser valores numéricos.')

        if not errores and t_max_f <= t_min_f:
            errores.append('La tarifa máxima debe ser mayor que la mínima.')

        if not errores and nombre != cat['nombre']:
            if db.execute('SELECT id FROM categorias WHERE nombre = ? AND id != ?', (nombre, cid)).fetchone():
                errores.append('Ya existe una categoría con ese nombre.')

        if errores:
            for e in errores:
                flash(e, 'error')
            return render_template(
                'admin/categorias/editar.html',
                seccion_activa='categorias',
                cat=cat,
                **_ctx()
            )

        db.execute(
            'UPDATE categorias SET nombre=?, descripcion=?, requisitos=?, tarifa_minima=?, tarifa_maxima=? WHERE id=?',
            (nombre, desc, requisitos or None, t_min_f, t_max_f, cid)
        )
        db.commit()
        flash(f'Categoría "{nombre}" actualizada correctamente.', 'success')
        return redirect(url_for('admin.categoria_detalle', cid=cid))

    return render_template(
        'admin/categorias/editar.html',
        seccion_activa='categorias',
        cat=cat,
        **_ctx()
    )


@admin_bp.route('/categorias/<int:cid>/desactivar', methods=['POST'])
def categoria_desactivar(cid):
    db  = get_db()
    cat = db.execute('SELECT * FROM categorias WHERE id = ?', (cid,)).fetchone()
    if not cat:
        flash('Categoría no encontrada.', 'error')
        return redirect(url_for('admin.categorias'))

    count = db.execute(
        "SELECT COUNT(*) FROM prestadores WHERE categoria_id = ? AND estado_perfil = 'APROBADO'",
        (cid,)
    ).fetchone()[0]

    if count > 0:
        flash(
            f'Esta categoría tiene {count} prestador{"es" if count != 1 else ""} '
            f'activo{"s" if count != 1 else ""}. Para desactivarla primero reasignales otra categoría.',
            'warning'
        )
        return redirect(url_for('admin.categoria_detalle', cid=cid))

    db.execute('UPDATE categorias SET activa = 0 WHERE id = ?', (cid,))
    db.commit()
    flash('Categoría desactivada.', 'success')
    return redirect(url_for('admin.categoria_detalle', cid=cid))


@admin_bp.route('/categorias/<int:cid>/activar', methods=['POST'])
def categoria_activar(cid):
    db  = get_db()
    cat = db.execute('SELECT * FROM categorias WHERE id = ?', (cid,)).fetchone()
    if not cat:
        flash('Categoría no encontrada.', 'error')
        return redirect(url_for('admin.categorias'))
    db.execute('UPDATE categorias SET activa = 1 WHERE id = ?', (cid,))
    db.commit()
    flash('Categoría activada correctamente.', 'success')
    return redirect(url_for('admin.categoria_detalle', cid=cid))


# ---------------------------------------------------------------------------
# Gestión de Pagos y Comisiones
# ---------------------------------------------------------------------------

def _get_comision_config(db):
    rows = {r['clave']: r['valor'] for r in db.execute(
        "SELECT clave, valor FROM configuracion "
        "WHERE clave IN ('comision_solicitante_pct','comision_prestador_pct')"
    ).fetchall()}
    sol_pct  = float(rows.get('comision_solicitante_pct', 15))
    pres_pct = float(rows.get('comision_prestador_pct', 7))
    return sol_pct, pres_pct


def _asegurar_config_comision(db):
    for clave, valor, desc in [
        ('comision_solicitante_pct', '15', 'Porcentaje que paga el solicitante sobre el monto del servicio'),
        ('comision_prestador_pct',   '7',  'Porcentaje que se descuenta al prestador sobre el monto del servicio'),
    ]:
        db.execute(
            'INSERT OR IGNORE INTO configuracion (clave, valor, descripcion) VALUES (?,?,?)',
            (clave, valor, desc)
        )


@admin_bp.route('/pagos')
def pagos():
    fecha_desde = request.args.get('fecha_desde', '')
    fecha_hasta = request.args.get('fecha_hasta', '')
    estado      = request.args.get('estado', '')
    metodo      = request.args.get('metodo_pago', '')
    busqueda    = request.args.get('busqueda', '').strip()

    db = get_db()
    query = '''
        SELECT pg.*,
               s.fecha_servicio,
               uf.nombre || ' ' || uf.apellido AS solicitante_nombre,
               up.nombre || ' ' || up.apellido AS prestador_nombre
        FROM pagos pg
        JOIN servicios   s  ON pg.servicio_id  = s.id
        JOIN solicitantes    f  ON pg.solicitante_id   = f.id
        JOIN usuarios    uf ON f.usuario_id    = uf.id
        JOIN prestadores pr ON pg.prestador_id = pr.id
        JOIN usuarios    up ON pr.usuario_id   = up.id
        WHERE 1=1
    '''
    params = []

    if fecha_desde:
        query += ' AND date(pg.fecha_pago) >= ?'
        params.append(fecha_desde)
    if fecha_hasta:
        query += ' AND date(pg.fecha_pago) <= ?'
        params.append(fecha_hasta)
    if estado:
        query += ' AND pg.estado = ?'
        params.append(estado)
    if metodo:
        query += ' AND pg.metodo_pago = ?'
        params.append(metodo)
    if busqueda:
        query += (' AND (uf.nombre LIKE ? OR uf.apellido LIKE ? '
                  'OR up.nombre LIKE ? OR up.apellido LIKE ?)')
        params += [f'%{busqueda}%'] * 4

    query += ' ORDER BY pg.id DESC'
    lista = db.execute(query, params).fetchall()

    total_bruto = sum(p['monto_bruto']    for p in lista)
    total_comis = sum(p['comision_monto'] for p in lista)
    total_neto  = sum(p['monto_neto']     for p in lista)

    url_pagos = {k[len('url_pago_'):]: session.get(k)
                 for k in list(session.keys()) if k.startswith('url_pago_')}

    return render_template(
        'admin/pagos/listado.html',
        seccion_activa='pagos',
        pagos=lista,
        total_bruto=total_bruto,
        total_comis=total_comis,
        total_neto=total_neto,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        estado=estado,
        metodo_pago=metodo,
        busqueda=busqueda,
        url_pagos=url_pagos,
        **_ctx()
    )


@admin_bp.route('/pagos/liquidaciones')
def pagos_liquidaciones():
    db = get_db()
    prestadores_pend = db.execute(
        '''SELECT pr.id AS prestador_id,
                  up.nombre || ' ' || up.apellido AS prestador_nombre,
                  c.nombre  AS categoria_nombre,
                  COUNT(pg.id)       AS cantidad_pagos,
                  SUM(pg.monto_neto) AS total_a_liquidar
           FROM pagos pg
           JOIN prestadores pr ON pg.prestador_id = pr.id
           JOIN usuarios    up ON pr.usuario_id   = up.id
           LEFT JOIN categorias c ON pr.categoria_id = c.id
           WHERE pg.estado = 'PROCESADO'
           GROUP BY pr.id
           ORDER BY up.apellido, up.nombre'''
    ).fetchall()

    return render_template(
        'admin/pagos/liquidaciones.html',
        seccion_activa='pagos',
        prestadores=prestadores_pend,
        **_ctx()
    )


@admin_bp.route('/pagos/liquidaciones/<int:prestador_id>/liquidar', methods=['POST'])
def pagos_liquidar_prestador(prestador_id):
    db = get_db()
    pagos_proc = db.execute(
        "SELECT id, monto_neto FROM pagos WHERE prestador_id=? AND estado='PROCESADO'",
        (prestador_id,)
    ).fetchall()

    if not pagos_proc:
        flash('No hay pagos procesados para liquidar.', 'error')
        return redirect(url_for('admin.pagos_liquidaciones'))

    total = sum(p['monto_neto'] for p in pagos_proc)
    ahora = ahora_argentina()

    for p in pagos_proc:
        db.execute(
            "UPDATE pagos SET estado='LIQUIDADO', fecha_liquidacion=? WHERE id=?",
            (ahora, p['id'])
        )

    pr = db.execute('SELECT usuario_id FROM prestadores WHERE id=?', (prestador_id,)).fetchone()
    if pr:
        _notificar(
            db, pr['usuario_id'], 'PAGO',
            f'Tu pago de ${total:,.0f} fue acreditado',
            f'Se liquidaron {len(pagos_proc)} pago(s) por un total de ${total:,.0f}.'
        )

    db.commit()
    if pr:
        u = db.execute('SELECT nombre, email FROM usuarios WHERE id=?', (pr['usuario_id'],)).fetchone()
        if u:
            mc_row = db.execute(
                'SELECT metodo_cobro_prestador FROM pagos WHERE id=?', (pagos_proc[0]['id'],)
            ).fetchone()
            metodo_cobro = (mc_row['metodo_cobro_prestador'] or '') if mc_row else ''
            asunto = _cfg_db('mail_pago_liquidado_asunto', 'Tu pago fue acreditado — AMPARO')
            cuerpo = _cfg_db('mail_pago_liquidado_cuerpo',
                'Hola {nombre},\n\nTu pago de {monto_neto} fue acreditado.\n\n'
                'Método de cobro: {metodo_cobro}\nFecha de liquidación: {fecha_liquidacion}\n\n'
                '{link_app}')
            cuerpo = (cuerpo
                .replace('{nombre}', u['nombre'])
                .replace('{monto_neto}', f'${total:,.0f}')
                .replace('{metodo_cobro}', metodo_cobro)
                .replace('{fecha_liquidacion}', ahora[:10])
                .replace('{link_app}', _cfg_db('app_url', 'http://127.0.0.1:5000') + '/prestador/dashboard')
            )
            enviar_email(u['email'], asunto, cuerpo)
    # Registrar movimientos financieros (uno por pago liquidado)
    try:
        from routes.financiero import registrar_movimiento
        pr_nombre_row = db.execute(
            'SELECT u.nombre, u.apellido FROM prestadores pr '
            'JOIN usuarios u ON u.id=pr.usuario_id WHERE pr.id=?',
            (prestador_id,)
        ).fetchone()
        pr_nombre = (f"{pr_nombre_row['nombre']} {pr_nombre_row['apellido']}"
                     if pr_nombre_row else f'prestador #{prestador_id}')
        for p in pagos_proc:
            registrar_movimiento(
                db, 'PAGO_PRESTADOR',
                f'Liquidación prestador {pr_nombre} — pago #{p["id"]}',
                monto_salida=p['monto_neto'],
                pago_id=p['id']
            )
        db.commit()
    except Exception as e:
        print(f'[AMPARO] Error registrando movimientos PAGO_PRESTADOR: {e}')
    flash(f'Se liquidaron {len(pagos_proc)} pago(s) por ${total:,.0f}.', 'success')
    return redirect(url_for('admin.pagos_liquidaciones'))


@admin_bp.route('/pagos/comisiones')
def pagos_comisiones():
    return redirect(url_for('admin.configuracion_comisiones'))


@admin_bp.route('/pagos/comisiones_legacy', methods=['GET', 'POST'])
def pagos_comisiones_legacy():
    db = get_db()
    _asegurar_config_comision(db)
    db.execute('''
        CREATE TABLE IF NOT EXISTS comision_historial (
            id         INTEGER  PRIMARY KEY AUTOINCREMENT,
            tipo_ant   TEXT,
            pct_ant    REAL,
            fijo_ant   REAL,
            tipo_nuevo TEXT,
            pct_nuevo  REAL,
            fijo_nuevo REAL,
            fecha      DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    db.commit()

    if request.method == 'POST':
        sol_s  = request.form.get('comision_sol_pct', '15').strip()
        pres_s = request.form.get('comision_pres_pct', '7').strip()

        errores = []
        sol_f = pres_f = None

        try:
            sol_f  = float(sol_s)
            pres_f = float(pres_s)
            if sol_f < 0 or sol_f > 100 or pres_f < 0 or pres_f > 100:
                errores.append('Los porcentajes deben estar entre 0 y 100.')
        except (ValueError, TypeError):
            errores.append('Los porcentajes deben ser números válidos.')

        if errores:
            for e in errores:
                flash(e, 'error')
        else:
            ahora = ahora_argentina()
            db.execute(
                "UPDATE configuracion SET valor=?, fecha_modificacion=? "
                "WHERE clave='comision_solicitante_pct'",
                (str(sol_f), ahora)
            )
            db.execute(
                "UPDATE configuracion SET valor=?, fecha_modificacion=? "
                "WHERE clave='comision_prestador_pct'",
                (str(pres_f), ahora)
            )
            db.commit()
            flash('Configuración de comisión actualizada correctamente.', 'success')
            return redirect(url_for('admin.pagos_comisiones_legacy'))

    sol_pct, pres_pct = _get_comision_config(db)
    historial = db.execute(
        'SELECT * FROM comision_historial ORDER BY fecha DESC LIMIT 5'
    ).fetchall()

    return render_template(
        'admin/pagos/configuracion_comisiones.html',
        seccion_activa='pagos',
        comision_tipo='porcentaje',
        comision_pct=sol_pct,
        comision_fijo=0,
        comision_sol_pct=sol_pct,
        comision_pres_pct=pres_pct,
        historial=historial,
        **_ctx()
    )


@admin_bp.route('/pagos/<int:pid>')
def pago_detalle(pid):
    db   = get_db()
    pago = db.execute(
        '''SELECT pg.*,
                  s.fecha_servicio, s.hora_inicio, s.hora_fin,
                  uf.nombre || ' ' || uf.apellido AS solicitante_nombre,
                  up.nombre || ' ' || up.apellido AS prestador_nombre,
                  up.id  AS prestador_usuario_id,
                  f.id   AS solicitante_fid,
                  pr.id  AS prestador_pid
           FROM pagos pg
           JOIN servicios   s  ON pg.servicio_id  = s.id
           JOIN solicitantes    f  ON pg.solicitante_id   = f.id
           JOIN usuarios    uf ON f.usuario_id    = uf.id
           JOIN prestadores pr ON pg.prestador_id = pr.id
           JOIN usuarios    up ON pr.usuario_id   = up.id
           WHERE pg.id = ?''',
        (pid,)
    ).fetchone()

    if not pago:
        flash('Pago no encontrado.', 'error')
        return redirect(url_for('admin.pagos'))

    return render_template(
        'admin/pagos/detalle.html',
        seccion_activa='pagos',
        pago=pago,
        url_pago=session.get(f'url_pago_{pid}'),
        **_ctx()
    )


@admin_bp.route('/pagos/<int:pid>/liquidar', methods=['POST'])
def pago_liquidar(pid):
    db   = get_db()
    pago = db.execute('SELECT * FROM pagos WHERE id=?', (pid,)).fetchone()
    if not pago or pago['estado'] != 'PROCESADO':
        flash('El pago no está en estado PROCESADO.', 'error')
        return redirect(url_for('admin.pago_detalle', pid=pid))

    ahora = ahora_argentina()
    db.execute(
        "UPDATE pagos SET estado='LIQUIDADO', fecha_liquidacion=? WHERE id=?",
        (ahora, pid)
    )
    pr = db.execute(
        'SELECT pr.usuario_id, u.nombre, u.apellido FROM prestadores pr '
        'JOIN usuarios u ON u.id=pr.usuario_id WHERE pr.id=?',
        (pago['prestador_id'],)
    ).fetchone()
    if pr:
        _notificar(
            db, pr['usuario_id'], 'PAGO',
            f'Tu pago de ${pago["monto_neto"]:,.0f} fue acreditado',
            'El pago correspondiente al servicio fue acreditado a tu cuenta.'
        )

    try:
        from routes.financiero import registrar_movimiento
        pr_nombre = (f"{pr['nombre']} {pr['apellido']}" if pr
                     else f'prestador #{pago["prestador_id"]}')
        registrar_movimiento(
            db, 'PAGO_PRESTADOR',
            f'Liquidación {pr_nombre} — servicio #{pago["servicio_id"]}',
            monto_salida=pago['monto_neto'],
            pago_id=pid
        )
    except Exception as e:
        print(f'[AMPARO] Error registrando PAGO_PRESTADOR: {e}')

    db.commit()
    flash('Liquidación confirmada correctamente.', 'success')
    return redirect(url_for('admin.pago_detalle', pid=pid))


@admin_bp.route('/pagos/<int:pid>/reembolsar', methods=['POST'])
def pago_reembolsar(pid):
    motivo = request.form.get('motivo', '').strip()
    if not motivo:
        flash('El motivo del reembolso es obligatorio.', 'error')
        return redirect(url_for('admin.pago_detalle', pid=pid))

    db   = get_db()
    pago = db.execute('SELECT * FROM pagos WHERE id=?', (pid,)).fetchone()
    if not pago or pago['estado'] not in ('PROCESADO', 'LIQUIDADO'):
        flash('Este pago no puede ser reembolsado en su estado actual.', 'error')
        return redirect(url_for('admin.pago_detalle', pid=pid))

    db.execute("UPDATE pagos SET estado='REEMBOLSADO' WHERE id=?", (pid,))

    fam = db.execute(
        'SELECT u.id FROM solicitantes f JOIN usuarios u ON f.usuario_id=u.id WHERE f.id=?',
        (pago['solicitante_id'],)
    ).fetchone()
    pr = db.execute('SELECT usuario_id FROM prestadores WHERE id=?',
                    (pago['prestador_id'],)).fetchone()
    monto_txt = f'${pago["monto_bruto"]:,.0f}'
    if fam:
        _notificar(db, fam['id'], 'PAGO',
                   f'Reembolso de {monto_txt} procesado', f'Motivo: {motivo}')
    if pr:
        _notificar(db, pr['usuario_id'], 'PAGO',
                   f'Reembolso de {monto_txt} procesado', f'Motivo: {motivo}')

    db.commit()
    flash('Reembolso iniciado correctamente.', 'success')
    return redirect(url_for('admin.pago_detalle', pid=pid))


@admin_bp.route('/pagos/<int:pid>/cobrar', methods=['POST'])
def pago_cobrar(pid):
    print(f"=== REINTENTAR COBRO pid={pid} ===")
    db   = get_db()
    pago = db.execute('SELECT * FROM pagos WHERE id=?', (pid,)).fetchone()

    if not pago:
        print(f"ERROR: Pago {pid} no encontrado en DB")
        flash('Pago no encontrado.', 'error')
        return redirect(url_for('admin.pagos'))

    print(f"Pago encontrado: id={pago['id']} estado={pago['estado']} monto_bruto={pago['monto_bruto']} referencia_pago={pago['referencia_pago']}")

    if pago['estado'] != 'PENDIENTE':
        print(f"ERROR: Estado es '{pago['estado']}', se requiere PENDIENTE")
        flash(f"Solo se puede reintentar cobro en pagos PENDIENTES (estado actual: {pago['estado']}).", 'error')
        return redirect(url_for('admin.pago_detalle', pid=pid))

    s = db.execute(
        """SELECT s.*, u.nombre||' '||u.apellido AS prestador_nombre
           FROM servicios s
           JOIN prestadores pr ON pr.id=s.prestador_id
           JOIN usuarios u ON u.id=pr.usuario_id
           WHERE s.id=?""",
        (pago['servicio_id'],)
    ).fetchone()
    print(f"Servicio: {dict(s) if s else 'NO ENCONTRADO'}")

    sol_row = db.execute(
        'SELECT uf.email FROM solicitantes f JOIN usuarios uf ON uf.id=f.usuario_id WHERE f.id=?',
        (pago['solicitante_id'],)
    ).fetchone()
    sol_email = sol_row['email'] if sol_row else 'pagos@amparo.com.ar'
    print(f"Email solicitante: {sol_email}")

    cfg_mp = {r['clave']: r['valor'] for r in db.execute(
        "SELECT clave, valor FROM configuracion WHERE clave IN ('mp_access_token','mp_modo','app_url')"
    ).fetchall()}
    access_token = cfg_mp.get('mp_access_token', '').strip()
    mp_modo      = cfg_mp.get('mp_modo', 'sandbox')
    app_url      = cfg_mp.get('app_url', '').strip()
    print(f"MP Token: {'[CONFIGURADO] ' + access_token[:20] + '...' if access_token else 'VACIO/NO CONFIGURADO'}")
    print(f"MP Modo: {mp_modo}")
    print(f"App URL en config: '{app_url}'")

    if not access_token:
        flash('No hay Access Token de Mercado Pago configurado. Configuralo en Admin > Configuracion > Pagos.', 'error')
        return redirect(url_for('admin.pago_detalle', pid=pid))

    try:
        import mercadopago
        print("Librería mercadopago importada OK")
        sdk = mercadopago.SDK(access_token)
        print("SDK MP inicializado")
        base_url = app_url or request.host_url.rstrip('/')
        print(f"base_url usado: {base_url}")
        prestador_nom = s['prestador_nombre'] if s else f'servicio #{pago["servicio_id"]}'
        preference_data = {
            "items": [{
                "title": f"Servicio AMPARO - {prestador_nom}",
                "quantity": 1,
                "unit_price": float(pago['monto_bruto']),
                "currency_id": "ARS",
            }],
            "payer": {"email": sol_email},
            "payment_methods": {"installments": 1},
            "external_reference": str(pid),
            "statement_descriptor": "AMPARO",
        }
        es_local = '127.0.0.1' in base_url or 'localhost' in base_url
        print(f"Entorno local: {es_local}")
        if not es_local:
            preference_data["back_urls"] = {
                "success": f"{base_url}/solicitante/pago/mp/ok?pago_id={pid}&sid={pago['servicio_id']}",
                "failure": f"{base_url}/solicitante/pago/mp/fallo?pago_id={pid}&sid={pago['servicio_id']}",
                "pending": f"{base_url}/solicitante/pago/mp/pendiente?pago_id={pid}&sid={pago['servicio_id']}",
            }
            preference_data["auto_return"]       = "approved"
            preference_data["notification_url"]  = f"{base_url}/solicitante/pago/mp/webhook"
        print(f"Creando preferencia MP con datos: {preference_data}")
        resp    = sdk.preference().create(preference_data)
        print(f"Respuesta MP completa: {resp}")
        pref    = resp.get("response", {})
        pref_id = pref.get("id")
        print(f"preference_id obtenido: {pref_id}")
        if pref_id:
            url_pago = pref.get('init_point') if mp_modo == 'produccion' else pref.get('sandbox_init_point')
            print(f"URL de pago: {url_pago}")
            db.execute("UPDATE pagos SET referencia_pago=? WHERE id=?", (pref_id, pid))
            db.commit()
            print(f"[OK] referencia_pago actualizado a {pref_id} para pago {pid}")
            session[f'url_pago_{pid}'] = url_pago
            flash(f'Preferencia MP creada. Usá el botón "Ver link de pago" para abrirla y simular el cobro.', 'success')
        else:
            print(f"ERROR: MP no devolvió id. pref={pref}")
            flash(f'MP no devolvió preference id. Respuesta: {pref}', 'error')
    except ImportError:
        print("ERROR CRÍTICO: librería 'mercadopago' no instalada")
        print("Ejecutar en terminal: pip install mercadopago")
        flash("Error: librería mercadopago no instalada. Ejecutar: pip install mercadopago", 'error')
    except Exception as e:
        import traceback
        print(f"ERROR en pago_cobrar: {e}")
        traceback.print_exc()
        flash(f'Error al crear preferencia MP: {str(e)[:200]}', 'error')

    return redirect(url_for('admin.pago_detalle', pid=pid))


@admin_bp.route('/pagos/<int:pid>/confirmar_pago', methods=['POST'])
def pago_confirmar_manual(pid):
    """Confirma un pago PENDIENTE manualmente (para dev local sin webhook)."""
    print(f"=== CONFIRMAR PAGO MANUAL pid={pid} ===")
    db   = get_db()
    pago = db.execute('SELECT * FROM pagos WHERE id=?', (pid,)).fetchone()
    if not pago or pago['estado'] != 'PENDIENTE':
        flash('Solo se puede confirmar manualmente un pago PENDIENTE.', 'error')
        return redirect(url_for('admin.pago_detalle', pid=pid))

    ahora = ahora_argentina()

    pr_row = db.execute(
        'SELECT usuario_id, metodo_cobro FROM prestadores WHERE id=?', (pago['prestador_id'],)
    ).fetchone()
    metodo_cobro_prestador = pr_row['metodo_cobro'] if pr_row else None

    db.execute(
        """UPDATE pagos SET estado='LIQUIDADO', metodo_pago='mercadopago',
           fecha_pago=?, fecha_liquidacion=?, metodo_cobro_prestador=? WHERE id=?""",
        (ahora, ahora, metodo_cobro_prestador, pid)
    )
    print(f"[OK] Pago {pid} marcado LIQUIDADO")

    if pr_row:
        _notificar(db, pr_row['usuario_id'], 'pago_liquidado',
                   f'Tu pago de $ {pago["monto_neto"]:,.0f} fue acreditado',
                   f'El pago de $ {pago["monto_neto"]:.2f} por tu servicio fue liquidado.')
    sol_row = db.execute(
        'SELECT uf.id FROM solicitantes f JOIN usuarios uf ON uf.id=f.usuario_id WHERE f.id=?',
        (pago['solicitante_id'],)
    ).fetchone()
    if sol_row:
        _notificar(db, sol_row['id'], 'pago_liquidado',
                   'Pago liquidado',
                   f'Tu pago de $ {pago["monto_bruto"]:.2f} fue procesado y liquidado.')
    admin_u = db.execute("SELECT id FROM usuarios WHERE tipo_usuario='admin' LIMIT 1").fetchone()
    if admin_u:
        _notificar(db, admin_u['id'], 'pago_liquidado',
                   'Pago liquidado',
                   f'Pago #{pid} — $ {pago["monto_bruto"]:.2f} liquidado.')

    # Registrar movimientos financieros (COBRO + PAGO_PRESTADOR en un solo paso)
    try:
        from routes.financiero import registrar_movimiento
        sol_row2 = db.execute(
            'SELECT u.nombre, u.apellido FROM solicitantes s '
            'JOIN usuarios u ON u.id=s.usuario_id WHERE s.id=?',
            (pago['solicitante_id'],)
        ).fetchone()
        sol_nombre = (f"{sol_row2['nombre']} {sol_row2['apellido']}" if sol_row2
                      else f'solicitante #{pago["solicitante_id"]}')
        total_cobrado = (pago['monto_bruto'] or 0) + (pago['comision_monto'] or 0)
        tipo_mov = 'PENALIDAD' if pago.get('tipo_pago') == 'penalidad' else 'COBRO'
        desc_cobro = (f"Penalidad cancelación — servicio #{pago['servicio_id']}"
                      if tipo_mov == 'PENALIDAD'
                      else f"Cobro servicio #{pago['servicio_id']} — {sol_nombre}")
        registrar_movimiento(
            db, tipo_mov, desc_cobro,
            monto_entrada=total_cobrado,
            pago_id=pid
        )
        pr_nombre2_row = db.execute(
            'SELECT u.nombre, u.apellido FROM prestadores pr '
            'JOIN usuarios u ON u.id=pr.usuario_id WHERE pr.id=?',
            (pago['prestador_id'],)
        ).fetchone()
        pr_nombre2 = (f"{pr_nombre2_row['nombre']} {pr_nombre2_row['apellido']}"
                      if pr_nombre2_row else f'prestador #{pago["prestador_id"]}')
        registrar_movimiento(
            db, 'PAGO_PRESTADOR',
            f'Liquidación {pr_nombre2} — servicio #{pago["servicio_id"]}',
            monto_salida=pago['monto_neto'],
            pago_id=pid
        )
    except Exception as e:
        print(f'[AMPARO] Error registrando movimientos en confirmación manual: {e}')

    db.commit()

    # Enviar correos con desglose
    try:
        from routes.prestador import _enviar_correos_liquidacion
        _enviar_correos_liquidacion(db, pid)
    except Exception as e:
        print(f"[AMPARO] Error enviando correos de liquidación: {e}")

    session.pop(f'url_pago_{pid}', None)
    flash('Pago liquidado correctamente.', 'success')
    return redirect(url_for('admin.pago_detalle', pid=pid))


# ═══════════════════════════════════════════════════════════════════════════════
# REPORTES Y ESTADÍSTICAS
# ═══════════════════════════════════════════════════════════════════════════════

def _periodo_fechas(req):
    """Retorna (periodo, desde_str, hasta_str) según parámetros del request."""
    from datetime import date, timedelta
    periodo = req.args.get('periodo', 'mes')
    hoy = date.today()
    if periodo == 'hoy':
        desde = hasta = hoy
    elif periodo == 'semana':
        desde = hoy - timedelta(days=hoy.weekday())
        hasta = hoy
    elif periodo == 'anio':
        desde = hoy.replace(month=1, day=1)
        hasta = hoy
    elif periodo == 'personalizado':
        try:
            desde = date.fromisoformat(req.args.get('desde', hoy.replace(day=1).isoformat()))
            hasta = date.fromisoformat(req.args.get('hasta', hoy.isoformat()))
        except ValueError:
            desde = hoy.replace(day=1)
            hasta = hoy
            periodo = 'mes'
    else:
        desde = hoy.replace(day=1)
        hasta = hoy
        periodo = 'mes'
    return periodo, desde.isoformat(), hasta.isoformat()


def _periodo_anterior(desde_str, hasta_str):
    from datetime import date, timedelta
    desde = date.fromisoformat(desde_str)
    hasta = date.fromisoformat(hasta_str)
    delta = (hasta - desde).days + 1
    p_hasta = desde - timedelta(days=1)
    p_desde = p_hasta - timedelta(days=delta - 1)
    return p_desde.isoformat(), p_hasta.isoformat()


def _label_periodo(periodo, desde, hasta):
    nombres = {'hoy': 'Hoy', 'semana': 'Esta semana', 'mes': 'Este mes',
               'anio': 'Este año', 'personalizado': f'{desde} al {hasta}'}
    return nombres.get(periodo, f'{desde} al {hasta}')


# ─── ÍNDICE DE REPORTES ───────────────────────────────────────────────────────

@admin_bp.route('/reportes')
def reportes():
    return render_template('admin/reportes/index.html',
                           seccion_activa='reportes', **_ctx())


# ─── REPORTE 1: USUARIOS ─────────────────────────────────────────────────────

@admin_bp.route('/reportes/usuarios')
def reporte_usuarios():
    db = get_db()
    periodo, desde, hasta = _periodo_fechas(request)
    p_desde, p_hasta = _periodo_anterior(desde, hasta)

    filas = db.execute('''
        SELECT strftime('%Y-%m', fecha_alta) AS mes,
               tipo_usuario, COUNT(*) AS cantidad
        FROM usuarios
        WHERE fecha_alta BETWEEN ? AND ?
        GROUP BY mes, tipo_usuario ORDER BY mes
    ''', (desde, hasta + ' 23:59:59')).fetchall()

    meses = sorted({r['mes'] for r in filas})
    tipos = ['admin', 'prestador', 'solicitante']
    chart_data = {t: [0] * len(meses) for t in tipos}
    for r in filas:
        # Contar admin_financiero junto con admin en los gráficos
        tipo_chart = 'admin' if r['tipo_usuario'] == 'admin_financiero' else r['tipo_usuario']
        if r['mes'] in meses and tipo_chart in tipos:
            chart_data[tipo_chart][meses.index(r['mes'])] += r['cantidad']

    tabla = []
    acum = 0
    for mes in meses:
        fila = {'mes': mes, 'admin': 0, 'prestador': 0, 'solicitante': 0}
        for r in filas:
            if r['mes'] == mes:
                tipo_key = 'admin' if r['tipo_usuario'] == 'admin_financiero' else r['tipo_usuario']
                if tipo_key in fila:
                    fila[tipo_key] += r['cantidad']
        fila['total'] = fila['admin'] + fila['prestador'] + fila['solicitante']
        acum += fila['total']
        fila['acumulado'] = acum
        tabla.append(fila)

    total_activos = db.execute(
        "SELECT COUNT(*) as c FROM usuarios WHERE fecha_bloqueo IS NULL"
    ).fetchone()['c']
    nuevos_periodo = sum(f['total'] for f in tabla)
    nuevos_anterior = db.execute(
        "SELECT COUNT(*) as c FROM usuarios WHERE fecha_alta BETWEEN ? AND ?",
        (p_desde, p_hasta + ' 23:59:59')
    ).fetchone()['c']
    crecimiento = round((nuevos_periodo - nuevos_anterior) / nuevos_anterior * 100, 1) \
        if nuevos_anterior > 0 else None

    return render_template('admin/reportes/usuarios.html',
                           periodo=periodo, desde=desde, hasta=hasta,
                           label_periodo=_label_periodo(periodo, desde, hasta),
                           meses=meses, chart_data=chart_data, tabla=tabla,
                           total_activos=total_activos,
                           nuevos_periodo=nuevos_periodo, crecimiento=crecimiento,
                           seccion_activa='reportes', **_ctx())


# ─── REPORTE 2: TRANSACCIONES ─────────────────────────────────────────────────

@admin_bp.route('/reportes/transacciones')
def reporte_transacciones():
    db = get_db()
    periodo, desde, hasta = _periodo_fechas(request)
    p_desde, p_hasta = _periodo_anterior(desde, hasta)

    filas = db.execute('''
        SELECT strftime('%Y-%m', fecha_pago) AS mes,
               COUNT(*) AS cantidad,
               SUM(monto_bruto) AS bruto, SUM(comision_monto) AS comisiones,
               SUM(monto_neto) AS neto
        FROM pagos WHERE estado = 'PAGADO' AND fecha_pago BETWEEN ? AND ?
        GROUP BY mes ORDER BY mes
    ''', (desde, hasta + ' 23:59:59')).fetchall()

    meses = [r['mes'] for r in filas]
    chart_montos = [round(r['bruto'] or 0, 2) for r in filas]

    metodos = db.execute('''
        SELECT COALESCE(metodo_pago, 'No especificado') AS metodo, COUNT(*) AS c
        FROM pagos WHERE estado = 'PAGADO' AND fecha_pago BETWEEN ? AND ?
        GROUP BY metodo
    ''', (desde, hasta + ' 23:59:59')).fetchall()

    tabla = []
    for r in filas:
        b = r['bruto'] or 0
        c = r['cantidad'] or 0
        tabla.append({
            'mes': r['mes'], 'cantidad': c, 'bruto': round(b, 2),
            'comisiones': round(r['comisiones'] or 0, 2),
            'neto': round(r['neto'] or 0, 2),
            'ticket_prom': round(b / c, 2) if c else 0,
        })

    total_tx    = sum(f['cantidad'] for f in tabla)
    total_bruto = sum(f['bruto'] for f in tabla)
    total_com   = sum(f['comisiones'] for f in tabla)
    ant = db.execute(
        "SELECT SUM(monto_bruto) as s FROM pagos WHERE estado='PAGADO' AND fecha_pago BETWEEN ? AND ?",
        (p_desde, p_hasta + ' 23:59:59')
    ).fetchone()['s'] or 0
    comparativa = round((total_bruto - ant) / ant * 100, 1) if ant else None

    return render_template('admin/reportes/transacciones.html',
                           periodo=periodo, desde=desde, hasta=hasta,
                           label_periodo=_label_periodo(periodo, desde, hasta),
                           meses=meses, chart_montos=chart_montos,
                           chart_metodos_labels=[r['metodo'] for r in metodos],
                           chart_metodos_data=[r['c'] for r in metodos],
                           tabla=tabla, total_tx=total_tx,
                           total_bruto=round(total_bruto, 2),
                           total_com=round(total_com, 2), comparativa=comparativa,
                           seccion_activa='reportes', **_ctx())


# ─── REPORTE 3: ZONAS ─────────────────────────────────────────────────────────

@admin_bp.route('/reportes/zonas')
def reporte_zonas():
    db = get_db()
    periodo, desde, hasta = _periodo_fechas(request)

    filas = db.execute('''
        SELECT z.nombre, z.ciudad, z.provincia,
               COUNT(s.id) AS solicitudes,
               SUM(CASE WHEN s.estado IN ('ACEPTADO','FINALIZADO') THEN 1 ELSE 0 END) AS contrataciones,
               SUM(CASE WHEN p.estado = 'PAGADO' THEN p.monto_bruto ELSE 0 END) AS facturado
        FROM zonas z
        LEFT JOIN solicitantes f ON f.zona_id = z.id
        LEFT JOIN servicios s ON s.solicitante_id = f.id
              AND s.fecha_solicitud BETWEEN ? AND ?
        LEFT JOIN pagos p ON p.servicio_id = s.id
        GROUP BY z.id ORDER BY solicitudes DESC
    ''', (desde, hasta + ' 23:59:59')).fetchall()

    tabla = []
    for i, r in enumerate(filas, 1):
        sol  = r['solicitudes'] or 0
        cont = r['contrataciones'] or 0
        tabla.append({
            'pos': i, 'nombre': r['nombre'],
            'ciudad': r['ciudad'] or '—', 'provincia': r['provincia'] or '—',
            'solicitudes': sol, 'contrataciones': cont,
            'conversion': round(cont / sol * 100, 1) if sol else 0,
            'facturado': round(r['facturado'] or 0, 2),
        })

    return render_template('admin/reportes/zonas.html',
                           periodo=periodo, desde=desde, hasta=hasta,
                           label_periodo=_label_periodo(periodo, desde, hasta),
                           tabla=tabla,
                           chart_labels=[r['nombre'] for r in tabla[:10]],
                           chart_data=[r['solicitudes'] for r in tabla[:10]],
                           seccion_activa='reportes', **_ctx())


# ─── REPORTE 4: CATEGORÍAS ────────────────────────────────────────────────────

@admin_bp.route('/reportes/categorias')
def reporte_categorias():
    db = get_db()
    periodo, desde, hasta = _periodo_fechas(request)

    filas = db.execute('''
        SELECT c.nombre,
               COUNT(DISTINCT pr.id) AS prestadores_activos,
               COUNT(s.id) AS servicios,
               SUM(CASE WHEN p.estado='PAGADO' THEN p.monto_bruto ELSE 0 END) AS facturado,
               AVG(cal.puntaje) AS puntaje_prom
        FROM categorias c
        LEFT JOIN prestadores pr ON pr.categoria_id = c.id
        LEFT JOIN servicios s ON s.categoria_id = c.id
              AND s.fecha_solicitud BETWEEN ? AND ?
        LEFT JOIN pagos p ON p.servicio_id = s.id
        LEFT JOIN calificaciones cal ON cal.servicio_id = s.id
        GROUP BY c.id ORDER BY servicios DESC
    ''', (desde, hasta + ' 23:59:59')).fetchall()

    tabla = []
    for r in filas:
        serv = r['servicios'] or 0
        fact = r['facturado'] or 0
        tabla.append({
            'nombre': r['nombre'],
            'prestadores_activos': r['prestadores_activos'] or 0,
            'servicios': serv, 'facturado': round(fact, 2),
            'ticket_prom': round(fact / serv, 2) if serv else 0,
            'puntaje_prom': round(r['puntaje_prom'] or 0, 1),
        })

    evo = db.execute('''
        SELECT strftime('%Y-%m', s.fecha_solicitud) AS mes,
               c.nombre AS categoria, COUNT(s.id) AS cantidad
        FROM servicios s JOIN categorias c ON c.id = s.categoria_id
        WHERE s.fecha_solicitud BETWEEN ? AND ?
        GROUP BY mes, c.id ORDER BY mes
    ''', (desde, hasta + ' 23:59:59')).fetchall()
    meses_evo = sorted({r['mes'] for r in evo})
    datasets_evo = {r['nombre']: [0] * len(meses_evo) for r in tabla}
    for r in evo:
        if r['mes'] in meses_evo and r['categoria'] in datasets_evo:
            datasets_evo[r['categoria']][meses_evo.index(r['mes'])] = r['cantidad']

    return render_template('admin/reportes/categorias.html',
                           periodo=periodo, desde=desde, hasta=hasta,
                           label_periodo=_label_periodo(periodo, desde, hasta),
                           tabla=tabla,
                           chart_torta_labels=[r['nombre'] for r in tabla],
                           chart_torta_data=[r['servicios'] for r in tabla],
                           meses_evo=meses_evo, datasets_evo=datasets_evo,
                           seccion_activa='reportes', **_ctx())


# ─── REPORTE 5: PRESTADORES ───────────────────────────────────────────────────

@admin_bp.route('/reportes/prestadores')
def reporte_prestadores():
    db = get_db()
    periodo, desde, hasta = _periodo_fechas(request)
    min_resenas = int(request.args.get('min_resenas', 3))
    cat_filtro  = request.args.get('categoria', '')
    categorias  = db.execute('SELECT id, nombre FROM categorias ORDER BY nombre').fetchall()

    where_cat = 'AND pr.categoria_id = ?' if cat_filtro else ''
    params = [desde, hasta + ' 23:59:59']
    if cat_filtro:
        params.append(int(cat_filtro))
    params.append(min_resenas)

    filas = db.execute(f'''
        SELECT u.nombre || ' ' || u.apellido AS prestador, pr.foto_url,
               c.nombre AS categoria, z.nombre AS zona,
               AVG(cal.puntaje) AS puntaje_prom, COUNT(cal.id) AS resenas,
               COUNT(DISTINCT s.id) AS servicios,
               SUM(CASE WHEN p.estado='PAGADO' THEN p.monto_bruto ELSE 0 END) AS facturado
        FROM prestadores pr
        JOIN usuarios u ON u.id = pr.usuario_id
        LEFT JOIN categorias c ON c.id = pr.categoria_id
        LEFT JOIN zonas z ON z.id = pr.zona_id
        LEFT JOIN servicios s ON s.prestador_id = pr.id
              AND s.fecha_solicitud BETWEEN ? AND ?
        LEFT JOIN calificaciones cal ON cal.prestador_id = pr.id
        LEFT JOIN pagos p ON p.servicio_id = s.id
        {where_cat}
        GROUP BY pr.id HAVING COUNT(cal.id) >= ?
        ORDER BY puntaje_prom DESC, resenas DESC LIMIT 50
    ''', params).fetchall()

    tabla = [{'pos': i, 'prestador': r['prestador'], 'foto_url': r['foto_url'],
              'categoria': r['categoria'] or '—', 'zona': r['zona'] or '—',
              'puntaje': round(r['puntaje_prom'] or 0, 1), 'resenas': r['resenas'],
              'servicios': r['servicios'], 'facturado': round(r['facturado'] or 0, 2)}
             for i, r in enumerate(filas, 1)]

    return render_template('admin/reportes/prestadores.html',
                           periodo=periodo, desde=desde, hasta=hasta,
                           label_periodo=_label_periodo(periodo, desde, hasta),
                           tabla=tabla, categorias=categorias,
                           cat_filtro=cat_filtro, min_resenas=min_resenas,
                           seccion_activa='reportes', **_ctx())


# ─── REPORTE 6: TIEMPOS ───────────────────────────────────────────────────────

@admin_bp.route('/reportes/tiempos')
def reporte_tiempos():
    db = get_db()
    periodo, desde, hasta = _periodo_fechas(request)

    filas = db.execute('''
        SELECT strftime('%Y-%m', fecha_solicitud) AS mes,
               COUNT(*) AS total,
               SUM(CASE WHEN estado IN ('ACEPTADO','FINALIZADO') THEN 1 ELSE 0 END) AS aceptados,
               SUM(CASE WHEN estado = 'RECHAZADO' THEN 1 ELSE 0 END) AS rechazados,
               AVG(CASE WHEN fecha_aceptacion IS NOT NULL
                   THEN (julianday(fecha_aceptacion) - julianday(fecha_solicitud)) * 24
                   END) AS hrs_sol_acept,
               AVG(CASE WHEN fecha_aceptacion IS NOT NULL AND fecha_servicio IS NOT NULL
                   THEN julianday(fecha_servicio) - julianday(fecha_aceptacion)
                   END) AS dias_acept_inicio
        FROM servicios
        WHERE fecha_solicitud BETWEEN ? AND ?
        GROUP BY mes ORDER BY mes
    ''', (desde, hasta + ' 23:59:59')).fetchall()

    tabla = []
    for r in filas:
        total = r['total'] or 0
        acept = r['aceptados'] or 0
        tabla.append({
            'mes': r['mes'], 'total': total, 'aceptados': acept,
            'rechazados': r['rechazados'] or 0,
            'tasa_acept': round(acept / total * 100, 1) if total else 0,
            'hrs_sol_acept': round(r['hrs_sol_acept'] or 0, 1),
            'dias_inicio': round(r['dias_acept_inicio'] or 0, 1),
        })

    return render_template('admin/reportes/tiempos.html',
                           periodo=periodo, desde=desde, hasta=hasta,
                           label_periodo=_label_periodo(periodo, desde, hasta),
                           tabla=tabla,
                           meses=[r['mes'] for r in tabla],
                           chart_hrs=[r['hrs_sol_acept'] for r in tabla],
                           chart_tasa=[r['tasa_acept'] for r in tabla],
                           seccion_activa='reportes', **_ctx())


# ─── EXPORTAR EXCEL ───────────────────────────────────────────────────────────

@admin_bp.route('/reportes/<reporte>/excel')
def reporte_excel(reporte):
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from flask import send_file

    VALIDOS = ['usuarios', 'transacciones', 'zonas', 'categorias', 'prestadores', 'tiempos']
    if reporte not in VALIDOS:
        return redirect(url_for('admin.reportes'))

    db = get_db()
    periodo, desde, hasta = _periodo_fechas(request)
    wb = Workbook()
    ws = wb.active
    AZUL = '1A5276'

    def _hrow(ws, cols, row=2):
        for i, col in enumerate(cols, 1):
            c = ws.cell(row=row, column=i, value=col)
            c.font = Font(bold=True, color='FFFFFF')
            c.fill = PatternFill('solid', fgColor=AZUL)
            c.alignment = Alignment(horizontal='center')

    def _titulo(ws, titulo, n):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n)
        c = ws.cell(row=1, column=1,
                    value=f'AMPARO — {titulo}  |  {_label_periodo(periodo, desde, hasta)}')
        c.font = Font(bold=True, size=13, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor=AZUL)
        c.alignment = Alignment(horizontal='center')
        ws.row_dimensions[1].height = 24

    if reporte == 'usuarios':
        ws.title = 'Usuarios'
        _titulo(ws, 'Reporte de Usuarios', 6)
        _hrow(ws, ['Mes', 'Admin', 'Prestadores', 'Familias', 'Total Nuevos', 'Acumulado'])
        rows_db = db.execute('''
            SELECT strftime('%Y-%m', fecha_alta) AS mes, tipo_usuario, COUNT(*) AS c
            FROM usuarios WHERE fecha_alta BETWEEN ? AND ?
            GROUP BY mes, tipo_usuario ORDER BY mes
        ''', (desde, hasta + ' 23:59:59')).fetchall()
        meses = sorted({r['mes'] for r in rows_db})
        acum = 0
        for mes in meses:
            d = {'admin': 0, 'prestador': 0, 'solicitante': 0}
            for r in rows_db:
                if r['mes'] == mes:
                    tipo_key = 'admin' if r['tipo_usuario'] == 'admin_financiero' else r['tipo_usuario']
                    if tipo_key in d:
                        d[tipo_key] += r['c']
            total = sum(d.values())
            acum += total
            ws.append([mes, d['admin'], d['prestador'], d['solicitante'], total, acum])

    elif reporte == 'transacciones':
        ws.title = 'Transacciones'
        _titulo(ws, 'Transacciones', 6)
        _hrow(ws, ['Mes', 'Cantidad', 'Monto Bruto', 'Comisiones', 'Monto Neto', 'Ticket Prom.'])
        for r in db.execute('''
            SELECT strftime('%Y-%m', fecha_pago) AS mes, COUNT(*) AS c,
                   SUM(monto_bruto) AS b, SUM(comision_monto) AS co, SUM(monto_neto) AS n
            FROM pagos WHERE estado='PAGADO' AND fecha_pago BETWEEN ? AND ?
            GROUP BY mes ORDER BY mes
        ''', (desde, hasta + ' 23:59:59')).fetchall():
            b, c = r['b'] or 0, r['c'] or 0
            ws.append([r['mes'], c, round(b,2), round(r['co'] or 0,2),
                       round(r['n'] or 0,2), round(b/c,2) if c else 0])

    elif reporte == 'zonas':
        ws.title = 'Zonas'
        _titulo(ws, 'Zonas con más demanda', 7)
        _hrow(ws, ['#','Zona','Ciudad','Provincia','Solicitudes','Contrataciones','Facturado'])
        for i, r in enumerate(db.execute('''
            SELECT z.nombre, z.ciudad, z.provincia,
                   COUNT(s.id) AS sol,
                   SUM(CASE WHEN s.estado IN ('ACEPTADO','FINALIZADO') THEN 1 ELSE 0 END) AS cont,
                   SUM(CASE WHEN p.estado='PAGADO' THEN p.monto_bruto ELSE 0 END) AS fact
            FROM zonas z
            LEFT JOIN solicitantes f ON f.zona_id = z.id
            LEFT JOIN servicios s ON s.solicitante_id = f.id AND s.fecha_solicitud BETWEEN ? AND ?
            LEFT JOIN pagos p ON p.servicio_id = s.id
            GROUP BY z.id ORDER BY sol DESC
        ''', (desde, hasta + ' 23:59:59')).fetchall(), 1):
            ws.append([i, r['nombre'], r['ciudad'] or '', r['provincia'] or '',
                       r['sol'] or 0, r['cont'] or 0, round(r['fact'] or 0,2)])

    elif reporte == 'categorias':
        ws.title = 'Categorías'
        _titulo(ws, 'Categorías más contratadas', 6)
        _hrow(ws, ['Categoría','Prestadores','Servicios','Facturado','Ticket Prom.','Calif. Prom.'])
        for r in db.execute('''
            SELECT c.nombre, COUNT(DISTINCT pr.id) AS pr_act, COUNT(s.id) AS serv,
                   SUM(CASE WHEN p.estado='PAGADO' THEN p.monto_bruto ELSE 0 END) AS fact,
                   AVG(cal.puntaje) AS punt
            FROM categorias c
            LEFT JOIN prestadores pr ON pr.categoria_id = c.id
            LEFT JOIN servicios s ON s.categoria_id = c.id AND s.fecha_solicitud BETWEEN ? AND ?
            LEFT JOIN pagos p ON p.servicio_id = s.id
            LEFT JOIN calificaciones cal ON cal.servicio_id = s.id
            GROUP BY c.id ORDER BY serv DESC
        ''', (desde, hasta + ' 23:59:59')).fetchall():
            s, f = r['serv'] or 0, r['fact'] or 0
            ws.append([r['nombre'], r['pr_act'] or 0, s, round(f,2),
                       round(f/s,2) if s else 0, round(r['punt'] or 0,1)])

    elif reporte == 'prestadores':
        ws.title = 'Prestadores'
        _titulo(ws, 'Prestadores mejor calificados', 6)
        _hrow(ws, ['#','Prestador','Categoría','Zona','Puntaje','Reseñas'])
        for i, r in enumerate(db.execute('''
            SELECT u.nombre || ' ' || u.apellido AS prest,
                   c.nombre AS cat, z.nombre AS zona,
                   AVG(cal.puntaje) AS punt, COUNT(cal.id) AS res
            FROM prestadores pr
            JOIN usuarios u ON u.id = pr.usuario_id
            LEFT JOIN categorias c ON c.id = pr.categoria_id
            LEFT JOIN zonas z ON z.id = pr.zona_id
            LEFT JOIN calificaciones cal ON cal.prestador_id = pr.id
            GROUP BY pr.id HAVING COUNT(cal.id) >= 1
            ORDER BY punt DESC LIMIT 50
        ''').fetchall(), 1):
            ws.append([i, r['prest'], r['cat'] or '', r['zona'] or '',
                       round(r['punt'] or 0,1), r['res']])

    else:  # tiempos
        ws.title = 'Tiempos'
        _titulo(ws, 'Tiempos de contratación', 5)
        _hrow(ws, ['Mes','Solicitudes','Tasa aceptación %','Hs. a aceptar','Días a inicio'])
        for r in db.execute('''
            SELECT strftime('%Y-%m', fecha_solicitud) AS mes, COUNT(*) AS total,
                   SUM(CASE WHEN estado IN ('ACEPTADO','FINALIZADO') THEN 1 ELSE 0 END) AS acept,
                   AVG(CASE WHEN fecha_aceptacion IS NOT NULL
                       THEN (julianday(fecha_aceptacion)-julianday(fecha_solicitud))*24 END) AS hrs,
                   AVG(CASE WHEN fecha_aceptacion IS NOT NULL AND fecha_servicio IS NOT NULL
                       THEN julianday(fecha_servicio)-julianday(fecha_aceptacion) END) AS dias
            FROM servicios WHERE fecha_solicitud BETWEEN ? AND ?
            GROUP BY mes ORDER BY mes
        ''', (desde, hasta + ' 23:59:59')).fetchall():
            t, a = r['total'] or 0, r['acept'] or 0
            ws.append([r['mes'], t, round(a/t*100,1) if t else 0,
                       round(r['hrs'] or 0,1), round(r['dias'] or 0,1)])

    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from flask import send_file
    return send_file(buf, as_attachment=True,
                     download_name=f'AMPARO_{reporte}_{desde}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ─── EXPORTAR PDF ─────────────────────────────────────────────────────────────

@admin_bp.route('/reportes/<reporte>/pdf')
def reporte_pdf(reporte):
    import io
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from flask import send_file
    from datetime import datetime as dt_now

    VALIDOS = ['usuarios', 'transacciones', 'zonas', 'categorias', 'prestadores', 'tiempos']
    if reporte not in VALIDOS:
        return redirect(url_for('admin.reportes'))

    db = get_db()
    periodo, desde, hasta = _periodo_fechas(request)

    TITULOS = {
        'usuarios': 'Usuarios Activos', 'transacciones': 'Transacciones',
        'zonas': 'Zonas con más demanda', 'categorias': 'Categorías más contratadas',
        'prestadores': 'Prestadores mejor calificados', 'tiempos': 'Tiempos de contratación',
    }
    AZUL_RL = colors.HexColor('#1A5276')

    if reporte == 'usuarios':
        headers = ['Mes', 'Admin', 'Prestadores', 'Familias', 'Total', 'Acumulado']
        rows_db = db.execute('''
            SELECT strftime('%Y-%m', fecha_alta) AS mes, tipo_usuario, COUNT(*) AS c
            FROM usuarios WHERE fecha_alta BETWEEN ? AND ?
            GROUP BY mes, tipo_usuario ORDER BY mes
        ''', (desde, hasta + ' 23:59:59')).fetchall()
        meses = sorted({r['mes'] for r in rows_db})
        data_rows, acum = [], 0
        for mes in meses:
            d = {'admin': 0, 'prestador': 0, 'solicitante': 0}
            for r in rows_db:
                if r['mes'] == mes:
                    tipo_key = 'admin' if r['tipo_usuario'] == 'admin_financiero' else r['tipo_usuario']
                    if tipo_key in d:
                        d[tipo_key] += r['c']
            total = sum(d.values())
            acum += total
            data_rows.append([mes, d['admin'], d['prestador'], d['solicitante'], total, acum])

    elif reporte == 'transacciones':
        headers = ['Mes', 'Cantidad', 'Monto Bruto', 'Comisiones', 'Neto', 'Ticket Prom.']
        data_rows = []
        for r in db.execute('''
            SELECT strftime('%Y-%m', fecha_pago) AS mes, COUNT(*) AS c,
                   SUM(monto_bruto) AS b, SUM(comision_monto) AS co, SUM(monto_neto) AS n
            FROM pagos WHERE estado='PAGADO' AND fecha_pago BETWEEN ? AND ?
            GROUP BY mes ORDER BY mes
        ''', (desde, hasta + ' 23:59:59')).fetchall():
            b, c = r['b'] or 0, r['c'] or 0
            data_rows.append([r['mes'], c, f'${b:.2f}', f'${r["co"] or 0:.2f}',
                               f'${r["n"] or 0:.2f}', f'${b/c:.2f}' if c else '$0'])

    elif reporte == 'zonas':
        headers = ['#', 'Zona', 'Ciudad', 'Solicitudes', 'Contrataciones', 'Facturado']
        data_rows = [[i+1, r['nombre'], r['ciudad'] or '', r['sol'] or 0,
                      r['cont'] or 0, f'${r["fact"] or 0:.2f}']
            for i, r in enumerate(db.execute('''
                SELECT z.nombre, z.ciudad,
                       COUNT(s.id) AS sol,
                       SUM(CASE WHEN s.estado IN ('ACEPTADO','FINALIZADO') THEN 1 ELSE 0 END) AS cont,
                       SUM(CASE WHEN p.estado='PAGADO' THEN p.monto_bruto ELSE 0 END) AS fact
                FROM zonas z
                LEFT JOIN solicitantes f ON f.zona_id = z.id
                LEFT JOIN servicios s ON s.solicitante_id = f.id AND s.fecha_solicitud BETWEEN ? AND ?
                LEFT JOIN pagos p ON p.servicio_id = s.id
                GROUP BY z.id ORDER BY sol DESC
            ''', (desde, hasta + ' 23:59:59')).fetchall())]

    elif reporte == 'categorias':
        headers = ['Categoría', 'Prestadores', 'Servicios', 'Facturado', 'Ticket', 'Calif.']
        data_rows = []
        for r in db.execute('''
            SELECT c.nombre, COUNT(DISTINCT pr.id) AS pr_act, COUNT(s.id) AS serv,
                   SUM(CASE WHEN p.estado='PAGADO' THEN p.monto_bruto ELSE 0 END) AS fact,
                   AVG(cal.puntaje) AS punt
            FROM categorias c
            LEFT JOIN prestadores pr ON pr.categoria_id = c.id
            LEFT JOIN servicios s ON s.categoria_id = c.id AND s.fecha_solicitud BETWEEN ? AND ?
            LEFT JOIN pagos p ON p.servicio_id = s.id
            LEFT JOIN calificaciones cal ON cal.servicio_id = s.id
            GROUP BY c.id ORDER BY serv DESC
        ''', (desde, hasta + ' 23:59:59')).fetchall():
            s, f = r['serv'] or 0, r['fact'] or 0
            data_rows.append([r['nombre'], r['pr_act'] or 0, s, f'${f:.2f}',
                               f'${f/s:.2f}' if s else '$0', f'{r["punt"] or 0:.1f}'])

    elif reporte == 'prestadores':
        headers = ['#', 'Prestador', 'Categoría', 'Zona', 'Puntaje', 'Reseñas']
        data_rows = [[i+1, r['prest'], r['cat'] or '', r['zona'] or '',
                      f'{r["punt"] or 0:.1f}', r['res']]
            for i, r in enumerate(db.execute('''
                SELECT u.nombre || ' ' || u.apellido AS prest,
                       c.nombre AS cat, z.nombre AS zona,
                       AVG(cal.puntaje) AS punt, COUNT(cal.id) AS res
                FROM prestadores pr
                JOIN usuarios u ON u.id = pr.usuario_id
                LEFT JOIN categorias c ON c.id = pr.categoria_id
                LEFT JOIN zonas z ON z.id = pr.zona_id
                LEFT JOIN calificaciones cal ON cal.prestador_id = pr.id
                GROUP BY pr.id HAVING COUNT(cal.id) >= 1
                ORDER BY punt DESC LIMIT 30
            ''').fetchall())]

    else:  # tiempos
        headers = ['Mes', 'Solicitudes', 'Aceptación %', 'Hs. a aceptar', 'Días a inicio']
        data_rows = []
        for r in db.execute('''
            SELECT strftime('%Y-%m', fecha_solicitud) AS mes, COUNT(*) AS total,
                   SUM(CASE WHEN estado IN ('ACEPTADO','FINALIZADO') THEN 1 ELSE 0 END) AS acept,
                   AVG(CASE WHEN fecha_aceptacion IS NOT NULL
                       THEN (julianday(fecha_aceptacion)-julianday(fecha_solicitud))*24 END) AS hrs,
                   AVG(CASE WHEN fecha_aceptacion IS NOT NULL AND fecha_servicio IS NOT NULL
                       THEN julianday(fecha_servicio)-julianday(fecha_aceptacion) END) AS dias
            FROM servicios WHERE fecha_solicitud BETWEEN ? AND ?
            GROUP BY mes ORDER BY mes
        ''', (desde, hasta + ' 23:59:59')).fetchall():
            t, a = r['total'] or 0, r['acept'] or 0
            data_rows.append([r['mes'], t, f'{a/t*100:.1f}%' if t else '0%',
                               f'{r["hrs"] or 0:.1f}', f'{r["dias"] or 0:.1f}'])

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    story = [
        Paragraph(f'<font color="#1A5276" size="16"><b>AMPARO — {TITULOS[reporte]}</b></font>',
                  styles['Normal']),
        Spacer(1, 0.3*cm),
        Paragraph(f'Período: {_label_periodo(periodo, desde, hasta)}  |  '
                  f'Generado: {dt_now.now().strftime("%d/%m/%Y %H:%M")}',
                  styles['Normal']),
        Spacer(1, 0.5*cm),
    ]
    table_data = [headers] + (data_rows or [['Sin datos en el período seleccionado']])
    t = Table(table_data, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND',    (0,0), (-1,0), AZUL_RL),
        ('TEXTCOLOR',     (0,0), (-1,0), colors.white),
        ('FONTNAME',      (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',      (0,0), (-1,-1), 8),
        ('ALIGN',         (0,0), (-1,-1), 'CENTER'),
        ('ROWBACKGROUNDS',(0,1), (-1,-1), [colors.white, colors.HexColor('#F0F4F8')]),
        ('GRID',          (0,0), (-1,-1), 0.5, colors.HexColor('#CCCCCC')),
        ('TOPPADDING',    (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.5*cm))
    story.append(Paragraph(
        f'<font size="7" color="grey">AMPARO — generado el {dt_now.now().strftime("%d/%m/%Y %H:%M")}</font>',
        styles['Normal']))
    doc.build(story)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f'AMPARO_{reporte}_{desde}.pdf',
                     mimetype='application/pdf')


# ─── Helpers de configuración ─────────────────────────────────────────────────

def _cfg(db, clave, default=''):
    row = db.execute('SELECT valor FROM configuracion WHERE clave=?', (clave,)).fetchone()
    return row['valor'] if row else default


def _set_cfg(db, clave, valor, descripcion=''):
    db.execute(
        '''INSERT INTO configuracion (clave, valor, descripcion, fecha_modificacion)
           VALUES (?, ?, ?, ?)
           ON CONFLICT(clave) DO UPDATE
           SET valor=excluded.valor, fecha_modificacion=excluded.fecha_modificacion''',
        (clave, valor, descripcion, datetime.now())
    )


# ─── CONFIGURACIÓN — ÍNDICE ──────────────────────────────────────────────────

@admin_bp.route('/configuracion')
def configuracion():
    return render_template('admin/configuracion/index.html',
                           seccion_activa='configuracion', **_ctx())


# ─── SECCIÓN 1: PASSWORDS ────────────────────────────────────────────────────

@admin_bp.route('/configuracion/passwords', methods=['GET', 'POST'])
def configuracion_passwords():
    db = get_db()
    if request.method == 'POST':
        vigencia = request.form.get('password_vigencia_dias', '').strip()
        longitud = request.form.get('password_min_longitud', '').strip()
        intentos = request.form.get('password_max_intentos', '').strip()
        errores = []
        if not vigencia.isdigit() or not (30 <= int(vigencia) <= 365):
            errores.append('La vigencia debe ser entre 30 y 365 días.')
        if not longitud.isdigit() or not (6 <= int(longitud) <= 20):
            errores.append('La longitud mínima debe ser entre 6 y 20 caracteres.')
        if not intentos.isdigit() or not (1 <= int(intentos) <= 10):
            errores.append('Los intentos deben ser entre 1 y 10.')
        if errores:
            for e in errores:
                flash(e, 'error')
        else:
            _set_cfg(db, 'password_vigencia_dias', vigencia)
            _set_cfg(db, 'password_min_longitud', longitud)
            _set_cfg(db, 'password_max_intentos', intentos)
            db.commit()
            flash('Configuración de contraseñas guardada correctamente.', 'success')
        return redirect(url_for('admin.configuracion_passwords'))

    config = {
        'password_vigencia_dias': _cfg(db, 'password_vigencia_dias', '90'),
        'password_min_longitud':  _cfg(db, 'password_min_longitud', '8'),
        'password_max_intentos':  _cfg(db, 'password_max_intentos', '3'),
    }
    return render_template('admin/configuracion/passwords.html',
                           config=config, seccion_activa='configuracion', **_ctx())


# ─── SECCIÓN 2: MAILS ────────────────────────────────────────────────────────

MAIL_CLAVES = [
    # (clave_base, label, variables_disponibles)
    ('bienvenida',                'Bienvenida (nuevo usuario)',           '{nombre} {link_app}'),
    ('desbloqueo',                'Desbloqueo de cuenta',                 '{nombre} {link_desbloqueo} {horas}'),
    ('contrasena_temp',           'Contraseña temporal (reseteo)',        '{nombre} {contrasena_temporal}'),
    ('vencimiento',               'Vencimiento próximo (10 días antes)',  '{nombre} {dias_restantes} {fecha_vencimiento}'),
    ('perfil_aprobado',           'Perfil aprobado (al prestador)',       '{nombre} {link_app}'),
    ('perfil_rechazado',          'Perfil rechazado (al prestador)',      '{nombre} {motivo_rechazo}'),
    ('registro_prestador',        'Registro recibido (prestador)',        '{nombre} {link_app}'),
    ('recibo_pago',               'Recibo de pago (solicitante)',         '{nombre} {prestador_nombre} {fecha_servicio} {hora_inicio} {hora_fin} {monto_servicio} {comision} {total_pagado}'),
    ('pago_liquidado',            'Pago liquidado (prestador)',           '{nombre} {monto_neto} {metodo_cobro} {fecha_liquidacion}'),
    ('cancelacion_sin_penalidad', 'Cancelación sin penalidad',           '{nombre} {prestador_nombre} {fecha_servicio} {hora_inicio} {hora_fin}'),
    ('cancelacion_con_penalidad', 'Cancelación con penalidad',           '{nombre} {prestador_nombre} {fecha_servicio} {hora_inicio} {hora_fin} {monto_penalidad}'),
    ('respuesta_contacto',        'Respuesta a mensaje de contacto',     '{nombre} {tipo_contacto} {asunto_mensaje} {respuesta_admin}'),
]


@admin_bp.route('/configuracion/mails', methods=['GET', 'POST'])
def configuracion_mails():
    db = get_db()
    if request.method == 'POST':
        for clave_base, _, _vars in MAIL_CLAVES:
            asunto = request.form.get(f'mail_{clave_base}_asunto', '').strip()
            cuerpo = request.form.get(f'mail_{clave_base}_cuerpo', '').strip()
            _set_cfg(db, f'mail_{clave_base}_asunto', asunto)
            _set_cfg(db, f'mail_{clave_base}_cuerpo', cuerpo)
        db.commit()
        flash('Textos de mails guardados correctamente.', 'success')
        return redirect(url_for('admin.configuracion_mails'))

    config = {}
    for clave_base, _, _vars in MAIL_CLAVES:
        config[f'mail_{clave_base}_asunto'] = _cfg(db, f'mail_{clave_base}_asunto', '')
        config[f'mail_{clave_base}_cuerpo'] = _cfg(db, f'mail_{clave_base}_cuerpo', '')
    return render_template('admin/configuracion/mails.html',
                           config=config, mail_claves=MAIL_CLAVES,
                           seccion_activa='configuracion', **_ctx())


# ─── SECCIÓN 3: EMPRESA ──────────────────────────────────────────────────────

_EMPRESA_CAMPOS = ['empresa_nombre', 'empresa_email', 'empresa_telefono',
                   'empresa_direccion', 'empresa_web', 'empresa_logo_url']


@admin_bp.route('/configuracion/empresa', methods=['GET', 'POST'])
def configuracion_empresa():
    db = get_db()
    if request.method == 'POST':
        for campo in ['empresa_nombre', 'empresa_email', 'empresa_telefono',
                      'empresa_direccion', 'empresa_web']:
            _set_cfg(db, campo, request.form.get(campo, '').strip())
        logo = request.files.get('logo')
        if logo and logo.filename:
            ext = logo.filename.rsplit('.', 1)[-1].lower()
            if ext in {'jpg', 'jpeg', 'png', 'webp', 'gif', 'svg'}:
                filename = f'logo_empresa.{ext}'
                logo_dir = os.path.join('static', 'img')
                os.makedirs(logo_dir, exist_ok=True)
                logo.save(os.path.join(logo_dir, filename))
                _set_cfg(db, 'empresa_logo_url', f'/static/img/{filename}')
            else:
                flash('Formato de imagen no válido.', 'error')
        db.commit()
        flash('Datos de la empresa guardados correctamente.', 'success')
        return redirect(url_for('admin.configuracion_empresa'))

    config = {c: _cfg(db, c, '') for c in _EMPRESA_CAMPOS}
    return render_template('admin/configuracion/empresa.html',
                           config=config, seccion_activa='configuracion', **_ctx())


# ─── SECCIÓN 4: ZONAS ────────────────────────────────────────────────────────

@admin_bp.route('/configuracion/zonas')
def configuracion_zonas():
    db = get_db()
    zonas = db.execute('''
        SELECT z.id, z.nombre, z.ciudad, z.provincia, z.activa,
               COUNT(p.id) AS cant_prestadores
        FROM zonas z
        LEFT JOIN prestadores p ON p.zona_id = z.id AND p.estado_perfil = 'APROBADO'
        GROUP BY z.id
        ORDER BY z.nombre
    ''').fetchall()
    return render_template('admin/configuracion/zonas.html',
                           zonas=zonas, seccion_activa='configuracion', **_ctx())


@admin_bp.route('/configuracion/zonas/nueva', methods=['GET', 'POST'])
def configuracion_zona_nueva():
    db = get_db()
    if request.method == 'POST':
        nombre = request.form.get('nombre', '').strip()
        ciudad = request.form.get('ciudad', '').strip()
        provincia = request.form.get('provincia', '').strip()
        if not nombre:
            flash('El nombre de la zona es obligatorio.', 'error')
            return render_template('admin/configuracion/zona_form.html',
                                   zona=None, seccion_activa='configuracion', **_ctx())
        if db.execute('SELECT id FROM zonas WHERE nombre=?', (nombre,)).fetchone():
            flash('Ya existe una zona con ese nombre.', 'error')
            return render_template('admin/configuracion/zona_form.html',
                                   zona=None, seccion_activa='configuracion', **_ctx())
        db.execute('INSERT INTO zonas (nombre, ciudad, provincia, activa) VALUES (?,?,?,1)',
                   (nombre, ciudad, provincia))
        db.commit()
        flash('Zona creada correctamente.', 'success')
        return redirect(url_for('admin.configuracion_zonas'))
    return render_template('admin/configuracion/zona_form.html',
                           zona=None, seccion_activa='configuracion', **_ctx())


@admin_bp.route('/configuracion/zonas/<int:zid>/editar', methods=['GET', 'POST'])
def configuracion_zona_editar(zid):
    db = get_db()
    zona = db.execute('SELECT * FROM zonas WHERE id=?', (zid,)).fetchone()
    if not zona:
        flash('Zona no encontrada.', 'error')
        return redirect(url_for('admin.configuracion_zonas'))
    if request.method == 'POST':
        nombre = request.form.get('nombre', '').strip()
        ciudad = request.form.get('ciudad', '').strip()
        provincia = request.form.get('provincia', '').strip()
        if not nombre:
            flash('El nombre es obligatorio.', 'error')
            return render_template('admin/configuracion/zona_form.html',
                                   zona=zona, seccion_activa='configuracion', **_ctx())
        if db.execute('SELECT id FROM zonas WHERE nombre=? AND id!=?', (nombre, zid)).fetchone():
            flash('Ya existe otra zona con ese nombre.', 'error')
            return render_template('admin/configuracion/zona_form.html',
                                   zona=zona, seccion_activa='configuracion', **_ctx())
        db.execute('UPDATE zonas SET nombre=?, ciudad=?, provincia=? WHERE id=?',
                   (nombre, ciudad, provincia, zid))
        db.commit()
        flash('Zona actualizada correctamente.', 'success')
        return redirect(url_for('admin.configuracion_zonas'))
    return render_template('admin/configuracion/zona_form.html',
                           zona=zona, seccion_activa='configuracion', **_ctx())


@admin_bp.route('/configuracion/zonas/<int:zid>/toggle', methods=['POST'])
def configuracion_zona_toggle(zid):
    db = get_db()
    zona = db.execute('SELECT * FROM zonas WHERE id=?', (zid,)).fetchone()
    if not zona:
        flash('Zona no encontrada.', 'error')
        return redirect(url_for('admin.configuracion_zonas'))
    if zona['activa']:
        cant = db.execute(
            'SELECT COUNT(*) as c FROM prestadores WHERE zona_id=? AND activo=1', (zid,)
        ).fetchone()['c']
        if cant > 0:
            flash(f'No se puede desactivar: la zona tiene {cant} prestador(es) activo(s).', 'error')
            return redirect(url_for('admin.configuracion_zonas'))
    nuevo_estado = 0 if zona['activa'] else 1
    db.execute('UPDATE zonas SET activa=? WHERE id=?', (nuevo_estado, zid))
    db.commit()
    flash(f'Zona {"activada" if nuevo_estado else "desactivada"} correctamente.', 'success')
    return redirect(url_for('admin.configuracion_zonas'))


# ─── SECCIÓN 5: NOTIFICACIONES ───────────────────────────────────────────────

NOTIF_EVENTOS = [
    ('notif_nueva_solicitud',       'Nueva solicitud de servicio recibida',        'Prestadores'),
    ('notif_servicio_cancelado',    'Servicio cancelado por el solicitante',            'Prestadores'),
    ('notif_pago_acreditado',       'Pago acreditado',                              'Prestadores'),
    ('notif_nueva_resena',          'Nueva reseña recibida',                        'Prestadores'),
    ('notif_documento_rechazado',   'Documento rechazado por el admin',             'Prestadores'),
    ('notif_perfil_aprobado',       'Perfil aprobado',                              'Prestadores'),
    ('notif_solicitud_aceptada',    'Solicitud aceptada por el prestador',          'Familias'),
    ('notif_solicitud_rechazada',   'Solicitud rechazada por el prestador',         'Familias'),
    ('notif_recordatorio_servicio', 'Recordatorio de servicio (X horas antes)',     'Familias'),
    ('notif_servicio_finalizado',   'Servicio finalizado — invitación a calificar', 'Familias'),
    ('notif_pago_procesado',        'Pago procesado',                               'Familias'),
    ('notif_reclamo_actualizado',   'Reclamo actualizado',                          'Familias'),
]


@admin_bp.route('/configuracion/notificaciones', methods=['GET', 'POST'])
def configuracion_notificaciones():
    db = get_db()
    if request.method == 'POST':
        for clave, _, _ in NOTIF_EVENTOS:
            _set_cfg(db, clave, '1' if request.form.get(clave) else '0')
        db.commit()
        flash('Configuración de notificaciones guardada correctamente.', 'success')
        return redirect(url_for('admin.configuracion_notificaciones'))

    config = {clave: _cfg(db, clave, '1') for clave, _, _ in NOTIF_EVENTOS}
    return render_template('admin/configuracion/notificaciones.html',
                           config=config, notif_eventos=NOTIF_EVENTOS,
                           seccion_activa='configuracion', **_ctx())


# ─── SECCIÓN 6: PAGOS ────────────────────────────────────────────────────────

@admin_bp.route('/configuracion/pagos', methods=['GET', 'POST'])
def configuracion_pagos_mp():
    db = get_db()
    if request.method == 'POST':
        public_key   = request.form.get('mp_public_key', '').strip()
        access_token = request.form.get('mp_access_token', '').strip()
        webhook_url  = request.form.get('mp_webhook_url', '').strip()
        modo         = request.form.get('mp_modo', 'sandbox')
        if modo not in ('sandbox', 'produccion'):
            modo = 'sandbox'
        _set_cfg(db, 'mp_public_key', public_key)
        if access_token and not set(access_token) <= {'*'}:
            _set_cfg(db, 'mp_access_token', access_token)
        _set_cfg(db, 'mp_webhook_url', webhook_url)
        _set_cfg(db, 'mp_modo', modo)
        db.commit()
        flash('Configuración de pagos guardada correctamente.', 'success')
        return redirect(url_for('admin.configuracion_pagos_mp'))

    token_real = _cfg(db, 'mp_access_token', '')
    token_mask = ('****' + token_real[-4:]) if len(token_real) > 4 else ('****' if token_real else '')
    config = {
        'mp_public_key':   _cfg(db, 'mp_public_key', ''),
        'mp_access_token': token_mask,
        'mp_webhook_url':  _cfg(db, 'mp_webhook_url', ''),
        'mp_modo':         _cfg(db, 'mp_modo', 'sandbox'),
    }
    return render_template('admin/configuracion/pagos.html',
                           config=config, seccion_activa='configuracion', **_ctx())


# ─── SECCIÓN: TARIFAS POR CATEGORÍA ─────────────────────────────────────────

@admin_bp.route('/configuracion/tarifas', methods=['GET', 'POST'])
def configuracion_tarifas():
    db = get_db()
    if request.method == 'POST':
        cat_id = request.form.get('categoria_id', '').strip()
        try:
            tarifa_min = float(request.form.get('tarifa_minima', '0').replace(',', '.'))
        except ValueError:
            tarifa_min = 0.0
        try:
            tarifa_max = float(request.form.get('tarifa_maxima', '0').replace(',', '.'))
        except ValueError:
            tarifa_max = 0.0
        db.execute('UPDATE categorias SET tarifa_minima=?, tarifa_maxima=? WHERE id=?',
                   (tarifa_min, tarifa_max, cat_id))
        db.commit()
        flash('Rango de tarifa actualizado correctamente.', 'success')
        return redirect(url_for('admin.configuracion_tarifas'))
    categorias = db.execute(
        'SELECT id, nombre, tarifa_minima, tarifa_maxima FROM categorias WHERE activa=1 ORDER BY nombre'
    ).fetchall()
    return render_template('admin/configuracion/tarifas.html',
                           categorias=categorias, seccion_activa='configuracion', **_ctx())


# ─── SECCIÓN 7: MANTENIMIENTO ────────────────────────────────────────────────

@admin_bp.route('/configuracion/mantenimiento')
def configuracion_mantenimiento():
    import sys
    import flask as flask_module
    db = get_db()
    db_path = os.path.normpath(os.path.join(os.path.dirname(__file__), '..', 'amparo.db'))
    db_size_mb = round(os.path.getsize(db_path) / (1024 * 1024), 2) if os.path.exists(db_path) else 0

    tablas = ['usuarios', 'prestadores', 'solicitantes', 'servicios', 'pagos',
              'calificaciones', 'reclamos', 'notificaciones', 'zonas', 'categorias']
    conteos = {}
    for t in tablas:
        try:
            conteos[t] = db.execute(f'SELECT COUNT(*) as c FROM {t}').fetchone()['c']
        except Exception:
            conteos[t] = '-'

    backup_dir = os.path.normpath(os.path.join(os.path.dirname(__file__), '..', 'backups'))
    backups = []
    if os.path.exists(backup_dir):
        for f in sorted(os.listdir(backup_dir), reverse=True):
            if f.endswith('.db'):
                fp   = os.path.join(backup_dir, f)
                size = os.path.getsize(fp)
                if size >= 1024 * 1024:
                    tamanio_str = f'{size / (1024 * 1024):.1f} MB'
                else:
                    tamanio_str = f'{size / 1024:.0f} KB'
                backups.append({
                    'nombre':  f,
                    'fecha':   datetime.fromtimestamp(os.path.getmtime(fp)).strftime('%d/%m/%Y %H:%M'),
                    'tamanio': tamanio_str,
                    'tipo':    'Seguridad' if 'PRE_RESTAURACION' in f else 'Normal',
                })

    estado = {
        'python_version':    sys.version.split()[0],
        'flask_version':     flask_module.__version__,
        'db_size_mb':        db_size_mb,
        'servidor_datetime': datetime.now().strftime('%d/%m/%Y %H:%M:%S'),
        'conteos':           conteos,
    }
    return render_template('admin/configuracion/mantenimiento.html',
                           estado=estado, backups=backups,
                           ultimo_backup=_cfg(db, 'ultimo_backup', 'Nunca'),
                           seccion_activa='configuracion', **_ctx())


@admin_bp.route('/configuracion/mantenimiento/backup', methods=['POST'])
def configuracion_backup():
    import shutil
    db = get_db()
    db_path    = os.path.normpath(os.path.join(os.path.dirname(__file__), '..', 'amparo.db'))
    backup_dir = os.path.normpath(os.path.join(os.path.dirname(__file__), '..', 'backups'))
    os.makedirs(backup_dir, exist_ok=True)
    ts   = datetime.now().strftime('%Y%m%d_%H%M%S')
    dest = os.path.join(backup_dir, f'amparo_{ts}.db')
    shutil.copy2(db_path, dest)
    _set_cfg(db, 'ultimo_backup', datetime.now().strftime('%d/%m/%Y %H:%M'))
    db.commit()
    flash(f'Backup generado: amparo_{ts}.db', 'success')
    return redirect(url_for('admin.configuracion_mantenimiento'))


@admin_bp.route('/configuracion/comisiones', methods=['GET', 'POST'])
def configuracion_comisiones():
    db = get_db()
    _asegurar_config_comision(db)
    # Ensure cancellation keys exist
    for clave, valor, desc in [
        ('cancelacion_penalidad_pct', '10', 'Porcentaje de penalidad por cancelación'),
        ('cancelacion_prestador_pct', '70', 'Del total de penalidad, % que va al prestador'),
        ('cancelacion_amparo_pct',    '30', 'Del total de penalidad, % que va a AMPARO'),
    ]:
        db.execute(
            'INSERT OR IGNORE INTO configuracion (clave, valor, descripcion) VALUES (?,?,?)',
            (clave, valor, desc)
        )
    db.commit()

    if request.method == 'POST':
        sol_s    = request.form.get('comision_sol_pct', '15').strip()
        pres_s   = request.form.get('comision_pres_pct', '7').strip()
        pen_s    = request.form.get('cancelacion_penalidad_pct', '10').strip()
        pen_pr_s = request.form.get('cancelacion_prestador_pct', '70').strip()
        pen_am_s = request.form.get('cancelacion_amparo_pct', '30').strip()

        errores = []

        try:
            sol_f  = float(sol_s)
            pres_f = float(pres_s)
            if sol_f < 0 or sol_f > 100 or pres_f < 0 or pres_f > 100:
                errores.append('Los porcentajes deben estar entre 0 y 100.')
        except (ValueError, TypeError):
            errores.append('Los porcentajes de comisión deben ser números válidos.')
            sol_f = pres_f = None

        try:
            pen_f    = float(pen_s)
            pen_pr_f = float(pen_pr_s)
            pen_am_f = float(pen_am_s)
            if pen_f < 0 or pen_f > 100:
                errores.append('La penalidad debe estar entre 0 y 100%.')
            elif abs(pen_pr_f + pen_am_f - 100) > 0.01 or pen_pr_f < 0 or pen_am_f < 0:
                errores.append('La distribución de la penalidad debe sumar 100% y ser positiva.')
        except (ValueError, TypeError):
            errores.append('Los valores de penalidad deben ser números válidos.')
            pen_f = pen_pr_f = pen_am_f = None

        if errores:
            for e in errores:
                flash(e, 'error')
        else:
            ahora = ahora_argentina()
            updates = [
                ('comision_solicitante_pct', str(sol_f)),
                ('comision_prestador_pct',   str(pres_f)),
                ('cancelacion_penalidad_pct', str(pen_f)),
                ('cancelacion_prestador_pct', str(pen_pr_f)),
                ('cancelacion_amparo_pct',    str(pen_am_f)),
            ]
            for clave, valor in updates:
                db.execute(
                    "UPDATE configuracion SET valor=?, fecha_modificacion=? WHERE clave=?",
                    (valor, ahora, clave)
                )
            db.commit()
            flash('✅ Comisiones actualizadas correctamente.', 'success')
            return redirect(url_for('admin.configuracion_comisiones'))

    cfg = {r['clave']: r['valor'] for r in db.execute(
        "SELECT clave, valor FROM configuracion WHERE clave IN "
        "('comision_solicitante_pct','comision_prestador_pct',"
        "'cancelacion_penalidad_pct','cancelacion_prestador_pct','cancelacion_amparo_pct')"
    ).fetchall()}

    return render_template(
        'admin/configuracion/comisiones.html',
        seccion_activa='configuracion',
        comision_sol_pct=float(cfg.get('comision_solicitante_pct', 15)),
        comision_pres_pct=float(cfg.get('comision_prestador_pct', 7)),
        cancelacion_penalidad_pct=float(cfg.get('cancelacion_penalidad_pct', 10)),
        cancelacion_prestador_pct=float(cfg.get('cancelacion_prestador_pct', 70)),
        cancelacion_amparo_pct=float(cfg.get('cancelacion_amparo_pct', 30)),
        **_ctx()
    )


@admin_bp.route('/configuracion/mantenimiento/backup/<nombre>/descargar')
def configuracion_backup_descargar(nombre):
    from flask import send_from_directory
    backup_dir = os.path.normpath(os.path.join(os.path.dirname(__file__), '..', 'backups'))
    return send_from_directory(backup_dir, os.path.basename(nombre), as_attachment=True)


@admin_bp.route('/configuracion/mantenimiento/restaurar', methods=['POST'])
def configuracion_restaurar():
    import shutil
    archivo      = request.form.get('archivo', '').strip()
    confirmacion = request.form.get('confirmacion', '').strip()

    if confirmacion != 'RESTAURAR':
        flash('Restauración cancelada. Debías escribir RESTAURAR.', 'warning')
        return redirect(url_for('admin.configuracion_mantenimiento'))

    # Sanitize: only allow known backup filenames (no path traversal)
    nombre_limpio = os.path.basename(archivo)
    backup_dir = os.path.normpath(os.path.join(os.path.dirname(__file__), '..', 'backups'))
    ruta_backup = os.path.join(backup_dir, nombre_limpio)
    db_path     = os.path.normpath(os.path.join(os.path.dirname(__file__), '..', 'amparo.db'))

    if not os.path.exists(ruta_backup):
        flash('Archivo de backup no encontrado.', 'error')
        return redirect(url_for('admin.configuracion_mantenimiento'))

    try:
        ts        = datetime.now().strftime('%Y%m%d_%H%M%S')
        seguridad = os.path.join(backup_dir, f'amparo_PRE_RESTAURACION_{ts}.db')
        shutil.copy2(db_path, seguridad)
        shutil.copy2(ruta_backup, db_path)
        flash(f'✅ Base de datos restaurada desde {nombre_limpio}. '
              f'Reiniciá la aplicación para aplicar los cambios.', 'success')
    except Exception as e:
        flash(f'Error al restaurar: {str(e)}', 'error')

    return redirect(url_for('admin.configuracion_mantenimiento'))


@admin_bp.route('/configuracion/mantenimiento/subir-restaurar', methods=['POST'])
def configuracion_subir_restaurar():
    import shutil
    from werkzeug.utils import secure_filename
    confirmacion = request.form.get('confirmacion', '').strip()
    archivo      = request.files.get('archivo_backup')

    if confirmacion != 'RESTAURAR':
        flash('Restauración cancelada. Debías escribir RESTAURAR.', 'warning')
        return redirect(url_for('admin.configuracion_mantenimiento'))

    if not archivo or not archivo.filename.endswith('.db'):
        flash('Solo se aceptan archivos .db', 'error')
        return redirect(url_for('admin.configuracion_mantenimiento'))

    backup_dir = os.path.normpath(os.path.join(os.path.dirname(__file__), '..', 'backups'))
    db_path    = os.path.normpath(os.path.join(os.path.dirname(__file__), '..', 'amparo.db'))
    os.makedirs(backup_dir, exist_ok=True)

    try:
        ts              = datetime.now().strftime('%Y%m%d_%H%M%S')
        nombre_guardado = f'amparo_subido_{ts}.db'
        ruta_guardado   = os.path.join(backup_dir, nombre_guardado)
        archivo.save(ruta_guardado)
        seguridad = os.path.join(backup_dir, f'amparo_PRE_RESTAURACION_{ts}.db')
        shutil.copy2(db_path, seguridad)
        shutil.copy2(ruta_guardado, db_path)
        flash(f'✅ Archivo subido y base de datos restaurada desde {secure_filename(archivo.filename)}. '
              f'Reiniciá la aplicación para aplicar los cambios.', 'success')
    except Exception as e:
        flash(f'Error al subir y restaurar: {str(e)}', 'error')

    return redirect(url_for('admin.configuracion_mantenimiento'))


# ─── SECCIÓN SMTP ────────────────────────────────────────────────────────────

_SMTP_CAMPOS = ['mail_servidor', 'mail_puerto', 'mail_usuario',
                'mail_password', 'mail_tls', 'mail_remitente', 'brevo_api_key']


@admin_bp.route('/configuracion/smtp', methods=['GET', 'POST'])
def configuracion_smtp():
    db = get_db()
    if request.method == 'POST':
        _set_cfg(db, 'mail_servidor',  request.form.get('mail_servidor', '').strip())
        _set_cfg(db, 'mail_puerto',    request.form.get('mail_puerto', '587').strip())
        _set_cfg(db, 'mail_usuario',   request.form.get('mail_usuario', '').strip())
        _set_cfg(db, 'mail_tls',       '1' if request.form.get('mail_tls') else '0')
        _set_cfg(db, 'mail_remitente', request.form.get('mail_remitente', '').strip())
        # Solo actualizar password si se ingresó uno nuevo (no vacío)
        nueva_pass = request.form.get('mail_password', '').strip()
        if nueva_pass:
            _set_cfg(db, 'mail_password', nueva_pass)
        # Solo actualizar API key si se ingresó una nueva
        nueva_api_key = request.form.get('brevo_api_key', '').strip()
        if nueva_api_key:
            _set_cfg(db, 'brevo_api_key', nueva_api_key)
        db.commit()
        flash('Configuración SMTP guardada correctamente.', 'success')
        return redirect(url_for('admin.configuracion_smtp'))

    config = {c: _cfg(db, c, '') for c in _SMTP_CAMPOS}
    return render_template('admin/configuracion/smtp.html',
                           config=config, seccion_activa='configuracion', **_ctx())


@admin_bp.route('/configuracion/smtp/prueba', methods=['POST'])
def configuracion_smtp_prueba():
    from auth import enviar_email
    db         = get_db()
    admin_user = db.execute('SELECT email FROM usuarios WHERE id=?',
                            (session['usuario_id'],)).fetchone()
    destino    = admin_user['email'] if admin_user else ''
    if not destino:
        flash('No se encontró el email del administrador.', 'error')
        return redirect(url_for('admin.configuracion_smtp'))

    try:
        enviar_email(
            destino,
            'AMPARO — Correo de prueba',
            'Este es un correo de prueba enviado desde la configuración SMTP de AMPARO.\n\n'
            'Si estás leyendo este mensaje, el envío de correos está funcionando correctamente.'
        )
        flash(f'Correo de prueba enviado correctamente a {destino}.', 'success')
    except Exception as e:
        flash(f'Error SMTP: {str(e)}', 'error')
        print(f"ERROR SMTP DETALLADO: {e}")
        import traceback
        traceback.print_exc()
    return redirect(url_for('admin.configuracion_smtp'))


@admin_bp.route('/reclamos')
def reclamos():
    return _construccion('reclamos', 'Reclamos')


# ---------------------------------------------------------------------------
# Contactos
# ---------------------------------------------------------------------------

@admin_bp.route('/contactos')
def contactos():
    db = get_db()
    tipo_f   = request.args.get('tipo', '')
    estado_f = request.args.get('estado', '')
    usuario_f = request.args.get('tipo_usuario', '')

    query  = '''SELECT c.*, u.nombre || ' ' || u.apellido AS usuario_nombre
                FROM contactos c
                JOIN usuarios u ON c.usuario_id = u.id
                WHERE 1=1'''
    params = []
    if tipo_f:
        query += ' AND c.tipo_contacto = ?'
        params.append(tipo_f)
    if estado_f:
        query += ' AND c.estado = ?'
        params.append(estado_f)
    if usuario_f:
        query += ' AND c.tipo_usuario = ?'
        params.append(usuario_f)
    query += ' ORDER BY c.fecha_envio DESC'

    lista = db.execute(query, params).fetchall()
    return render_template('admin/contactos/listado.html',
                           seccion_activa='contactos',
                           contactos=lista,
                           tipo_f=tipo_f, estado_f=estado_f, usuario_f=usuario_f,
                           **_ctx())


@admin_bp.route('/contactos/<int:cid>')
def contacto_detalle(cid):
    db = get_db()
    c  = db.execute(
        '''SELECT c.*, u.nombre || ' ' || u.apellido AS usuario_nombre, u.email AS usuario_email
           FROM contactos c JOIN usuarios u ON c.usuario_id = u.id
           WHERE c.id = ?''', (cid,)
    ).fetchone()
    if not c:
        flash('Contacto no encontrado.', 'error')
        return redirect(url_for('admin.contactos'))
    # Marcar como EN_REVISION si era NUEVO
    if c['estado'] == 'NUEVO':
        db.execute("UPDATE contactos SET estado='EN_REVISION' WHERE id=?", (cid,))
        db.commit()
    return render_template('admin/contactos/detalle.html',
                           seccion_activa='contactos',
                           c=c, **_ctx())


@admin_bp.route('/contactos/<int:cid>/resolver', methods=['POST'])
def contacto_resolver(cid):
    db       = get_db()
    respuesta = request.form.get('respuesta', '').strip()
    c = db.execute('SELECT * FROM contactos WHERE id=?', (cid,)).fetchone()
    if not c:
        flash('Contacto no encontrado.', 'error')
        return redirect(url_for('admin.contactos'))
    db.execute(
        "UPDATE contactos SET estado='RESUELTO', respuesta=?, fecha_resolucion=? WHERE id=?",
        (respuesta or None, datetime.now().isoformat(), cid)
    )
    tipos_label = {'problema_tecnico': 'problema técnico', 'reclamo': 'reclamo', 'sugerencia': 'sugerencia'}
    tipo_str = tipos_label.get(c['tipo_contacto'], 'mensaje')
    _notificar(db, c['usuario_id'], 'contacto',
               f"Tenes una respuesta de AMPARO",
               f"Respondimos tu {tipo_str}: '{c['asunto']}'. Ingresa a Contacto para leerla.")
    db.commit()
    if c['usuario_id']:
        u = db.execute('SELECT nombre, email, tipo_usuario FROM usuarios WHERE id=?', (c['usuario_id'],)).fetchone()
        if u:
            _base = _cfg_db('app_url', 'http://127.0.0.1:5000')
            _link_contacto = _base + '/login'
            asunto_mail = _cfg_db('mail_respuesta_contacto_asunto', 'Respuesta a tu mensaje — AMPARO')
            cuerpo_mail = _cfg_db('mail_respuesta_contacto_cuerpo',
                'Hola {nombre},\n\nRespondimos tu {tipo_contacto}: "{asunto_mensaje}".\n\n'
                '{respuesta_admin}\n\n{link_app}')
            cuerpo_mail = (cuerpo_mail
                .replace('{nombre}', u['nombre'])
                .replace('{tipo_contacto}', tipo_str)
                .replace('{asunto_mensaje}', c['asunto'] or '')
                .replace('{respuesta_admin}', respuesta or '')
                .replace('{link_app}', _link_contacto)
            )
            enviar_email(u['email'], asunto_mail, cuerpo_mail)
    flash('Contacto marcado como resuelto.', 'success')
    return redirect(url_for('admin.contacto_detalle', cid=cid))


# ---------------------------------------------------------------------------
# Conflictos
# ---------------------------------------------------------------------------

@admin_bp.route('/conflictos')
def conflictos():
    db   = get_db()
    rows = db.execute(
        """SELECT s.*,
                  up.nombre || ' ' || up.apellido AS prestador_nombre,
                  uf.nombre || ' ' || uf.apellido AS solicitante_nombre
           FROM servicios s
           JOIN prestadores pr ON pr.id = s.prestador_id
           JOIN usuarios up ON up.id = pr.usuario_id
           JOIN solicitantes fam ON fam.id = s.solicitante_id
           JOIN usuarios uf ON uf.id = fam.usuario_id
           WHERE s.conflicto = 1 AND s.estado = 'ACTIVO'
           ORDER BY s.fecha_servicio DESC"""
    ).fetchall()
    return render_template('admin/servicios/conflictos.html',
                           rows=rows, seccion_activa='conflictos', **_ctx())


@admin_bp.route('/conflictos/<int:sid>')
def conflicto_detalle(sid):
    db = get_db()
    s  = db.execute(
        """SELECT s.*,
                  up.nombre || ' ' || up.apellido AS prestador_nombre,
                  up.email AS prestador_email, up.telefono AS prestador_telefono,
                  uf.nombre || ' ' || uf.apellido AS solicitante_nombre,
                  uf.email AS solicitante_email, uf.telefono AS solicitante_telefono,
                  c.nombre AS categoria_nombre
           FROM servicios s
           JOIN prestadores pr ON pr.id = s.prestador_id
           JOIN usuarios up ON up.id = pr.usuario_id
           JOIN solicitantes fam ON fam.id = s.solicitante_id
           JOIN usuarios uf ON uf.id = fam.usuario_id
           LEFT JOIN categorias c ON c.id = s.categoria_id
           WHERE s.id=?""",
        (sid,)
    ).fetchone()
    if not s:
        flash('Servicio no encontrado.', 'error')
        return redirect(url_for('admin.conflictos'))
    return render_template('admin/servicios/conflicto_detalle.html',
                           s=s, seccion_activa='conflictos', **_ctx())


@admin_bp.route('/conflictos/<int:sid>/resolver', methods=['POST'])
def conflicto_resolver(sid):
    db      = get_db()
    accion  = request.form.get('accion', '')
    nota    = request.form.get('nota', '').strip()

    s = db.execute("SELECT * FROM servicios WHERE id=?", (sid,)).fetchone()
    if not s:
        flash('Servicio no encontrado.', 'error')
        return redirect(url_for('admin.conflictos'))

    def _notif(uid, tipo, titulo, msg):
        db.execute(
            'INSERT INTO notificaciones (usuario_id, tipo, titulo, mensaje) VALUES (?,?,?,?)',
            (uid, tipo, titulo, msg)
        )

    pr_uid  = db.execute('SELECT usuario_id FROM prestadores WHERE id=?', (s['prestador_id'],)).fetchone()
    fam_uid = db.execute('SELECT usuario_id FROM solicitantes WHERE id=?', (s['solicitante_id'],)).fetchone()

    if accion == 'finalizar':
        # Calcular y crear pago igual que en el prestador
        cfg = {r['clave']: r['valor'] for r in db.execute(
            "SELECT clave, valor FROM configuracion "
            "WHERE clave IN ('comision_solicitante_pct','comision_prestador_pct')"
        ).fetchall()}
        try:
            horas = _calcular_horas(s['hora_inicio'], s['hora_fin'])
        except Exception:
            horas = 0
        tarifa_hora = s['tarifa_hora'] or 0
        monto_bruto = round(tarifa_hora * horas, 2) if (tarifa_hora and horas) else (s['monto_estimado'] or 0)
        sol_pct              = float(cfg.get('comision_solicitante_pct', 15))
        pres_pct             = float(cfg.get('comision_prestador_pct', 7))
        comision_solicitante = round(monto_bruto * sol_pct  / 100, 2)
        comision_prestador   = round(monto_bruto * pres_pct / 100, 2)
        comision_monto       = round(comision_solicitante + comision_prestador, 2)
        comision_pct         = 0
        monto_neto           = round(monto_bruto - comision_prestador, 2)

        db.execute("UPDATE servicios SET estado='FINALIZADO', fecha_finalizacion=? WHERE id=?",
                   (ahora_argentina(), sid))
        db.execute(
            """INSERT INTO pagos
               (servicio_id, solicitante_id, prestador_id, tipo_pago,
                monto_bruto, comision_pct, comision_monto,
                comision_solicitante, comision_prestador,
                monto_neto, estado)
               VALUES (?,?,?,'servicio',?,?,?,?,?,?,'PENDIENTE')""",
            (sid, s['solicitante_id'], s['prestador_id'],
             monto_bruto, comision_pct, comision_monto,
             comision_solicitante, comision_prestador, monto_neto)
        )
        msg_base = nota or 'El conflicto fue resuelto por el equipo de AMPARO. El servicio fue confirmado.'
        if pr_uid:
            _notif(pr_uid['usuario_id'], 'conflicto_resuelto',
                   'Conflicto resuelto — servicio confirmado', msg_base)
        if fam_uid:
            _notif(fam_uid['usuario_id'], 'conflicto_resuelto',
                   'Conflicto resuelto — servicio confirmado',
                   msg_base + f' Por favor completá el pago de $ {monto_bruto:.2f}.')
        flash('Servicio marcado como FINALIZADO. Se generó el cobro.', 'success')

    elif accion == 'cancelar':
        db.execute("UPDATE servicios SET estado='CANCELADO', motivo_cancelacion=? WHERE id=?",
                   (nota or 'Cancelado por AMPARO tras revisión del conflicto.', sid))
        msg_base = nota or 'El equipo de AMPARO revisó el caso y el servicio fue anulado.'
        if pr_uid:
            _notif(pr_uid['usuario_id'], 'conflicto_resuelto',
                   'Conflicto resuelto — servicio anulado', msg_base)
        if fam_uid:
            _notif(fam_uid['usuario_id'], 'conflicto_resuelto',
                   'Conflicto resuelto — servicio anulado', msg_base)
        flash('Servicio marcado como CANCELADO.', 'success')
    else:
        flash('Acción no reconocida.', 'error')
        return redirect(url_for('admin.conflicto_detalle', sid=sid))

    db.commit()
    return redirect(url_for('admin.conflictos'))


# ---------------------------------------------------------------------------
# Monitoreo de Servicios
# ---------------------------------------------------------------------------

@admin_bp.route('/servicios')
def servicios_admin():
    db = get_db()

    fecha  = request.args.get('fecha', '').strip()
    estado = request.args.get('estado', '').strip()
    busq   = request.args.get('busqueda', '').strip()

    q = '''SELECT s.id, s.fecha_servicio, s.hora_inicio, s.hora_fin, s.estado,
                  s.prestador_confirmo_llegada, s.prestador_confirmo_fin,
                  s.solicitante_confirmo_fin, s.conflicto,
                  pu.nombre || ' ' || pu.apellido AS prestador_nombre,
                  fu.nombre || ' ' || fu.apellido AS solicitante_nombre,
                  c.nombre AS categoria_nombre
           FROM servicios s
           JOIN prestadores p ON s.prestador_id = p.id
           JOIN usuarios pu   ON p.usuario_id   = pu.id
           JOIN solicitantes  f   ON s.solicitante_id   = f.id
           JOIN usuarios fu   ON f.usuario_id   = fu.id
           LEFT JOIN categorias c ON p.categoria_id = c.id
           WHERE 1=1'''
    params = []

    if fecha:
        q += ' AND s.fecha_servicio = ?'
        params.append(fecha)
    if estado:
        q += ' AND s.estado = ?'
        params.append(estado)
    if busq:
        like = f'%{busq}%'
        q += ' AND (pu.nombre || " " || pu.apellido LIKE ? OR fu.nombre || " " || fu.apellido LIKE ?)'
        params += [like, like]

    q += ' ORDER BY s.fecha_servicio DESC, s.hora_inicio DESC LIMIT 100'

    servicios = db.execute(q, params).fetchall()

    return render_template(
        'admin/servicios/index.html',
        seccion_activa='servicios',
        servicios=servicios,
        fecha=fecha,
        estado=estado,
        busqueda=busq,
        **_ctx()
    )


@admin_bp.route('/servicios/<int:sid>')
def servicio_admin_detalle(sid):
    db = get_db()

    s = db.execute(
        '''SELECT s.*,
                  pu.nombre || ' ' || pu.apellido AS prestador_nombre,
                  pu.email AS prestador_email, pu.telefono AS prestador_telefono,
                  fu.nombre || ' ' || fu.apellido AS solicitante_nombre,
                  fu.email AS solicitante_email, fu.telefono AS solicitante_telefono,
                  c.nombre AS categoria_nombre,
                  f.direccion, f.latitud AS solicitante_lat, f.longitud AS solicitante_lon
           FROM servicios s
           JOIN prestadores p ON s.prestador_id = p.id
           JOIN usuarios pu   ON p.usuario_id   = pu.id
           JOIN solicitantes  f   ON s.solicitante_id   = f.id
           JOIN usuarios fu   ON f.usuario_id   = fu.id
           LEFT JOIN categorias c ON p.categoria_id = c.id
           WHERE s.id = ?''',
        (sid,)
    ).fetchone()

    if not s:
        flash('Servicio no encontrado.', 'error')
        return redirect(url_for('admin.servicios_admin'))

    pago = db.execute(
        'SELECT * FROM pagos WHERE servicio_id=? ORDER BY id DESC LIMIT 1', (sid,)
    ).fetchone()

    notifs = db.execute(
        '''SELECT n.fecha, n.tipo, n.titulo, n.mensaje,
                  u.nombre || ' ' || u.apellido AS destinatario
           FROM notificaciones n
           JOIN usuarios u ON n.usuario_id = u.id
           WHERE n.titulo LIKE ? OR n.mensaje LIKE ?
           ORDER BY n.fecha DESC LIMIT 10''',
        (f'%#{sid}%', f'%#{sid}%')
    ).fetchall()

    return render_template(
        'admin/servicios/detalle.html',
        seccion_activa='servicios',
        s=s,
        pago=pago,
        notifs=notifs,
        **_ctx()
    )


@admin_bp.route('/servicios/<int:sid>/finalizar-manual', methods=['POST'])
def servicio_finalizar_manual(sid):
    db   = get_db()
    nota = request.form.get('nota', '').strip()

    s = db.execute(
        '''SELECT s.*,
                  p.tarifa_hora, p.monto_estimado,
                  p.usuario_id AS pr_usuario_id
           FROM servicios s
           JOIN prestadores p ON s.prestador_id = p.id
           WHERE s.id = ? AND s.estado = 'ACTIVO'
             AND s.prestador_confirmo_fin = 1
             AND s.solicitante_confirmo_fin = 0''',
        (sid,)
    ).fetchone()

    if not s:
        flash('No se puede finalizar este servicio manualmente.', 'error')
        return redirect(url_for('admin.servicio_admin_detalle', sid=sid))

    cfg = {r['clave']: r['valor'] for r in db.execute(
        "SELECT clave, valor FROM configuracion "
        "WHERE clave IN ('comision_solicitante_pct','comision_prestador_pct')"
    ).fetchall()}

    try:
        horas = _calcular_horas(s['hora_inicio'], s['hora_fin'])
    except Exception:
        horas = 0

    tarifa_hora          = s['tarifa_hora'] or 0
    monto_bruto          = round(tarifa_hora * horas, 2) if (tarifa_hora and horas) else (s['monto_estimado'] or 0)
    sol_pct              = float(cfg.get('comision_solicitante_pct', 15))
    pres_pct             = float(cfg.get('comision_prestador_pct', 7))
    comision_solicitante = round(monto_bruto * sol_pct  / 100, 2)
    comision_prestador   = round(monto_bruto * pres_pct / 100, 2)
    comision_monto       = round(comision_solicitante + comision_prestador, 2)
    comision_pct         = 0
    monto_neto           = round(monto_bruto - comision_prestador, 2)

    db.execute(
        "UPDATE servicios SET estado='FINALIZADO', solicitante_confirmo_fin=1, "
        "fecha_confirmacion_solicitante=? WHERE id=?",
        (ahora_argentina(), sid)
    )
    db.execute(
        """INSERT INTO pagos
           (servicio_id, solicitante_id, prestador_id, tipo_pago,
            monto_bruto, comision_pct, comision_monto,
            comision_solicitante, comision_prestador,
            monto_neto, estado)
           VALUES (?,?,?,'servicio',?,?,?,?,?,?,'PENDIENTE')""",
        (sid, s['solicitante_id'], s['prestador_id'],
         monto_bruto, comision_pct, comision_monto,
         comision_solicitante, comision_prestador, monto_neto)
    )

    msg = nota or 'El servicio fue finalizado por el equipo de AMPARO.'
    _notificar(db, s['pr_usuario_id'], 'servicio_finalizado',
               'Servicio finalizado manualmente', msg)
    fam_uid = db.execute('SELECT usuario_id FROM solicitantes WHERE id=?', (s['solicitante_id'],)).fetchone()
    if fam_uid:
        _notificar(db, fam_uid['usuario_id'], 'servicio_finalizado',
                   'Servicio finalizado',
                   msg + f' Por favor completá el pago de $ {monto_bruto:.2f}.')
    db.commit()
    flash('Servicio finalizado manualmente. Se generó el cobro.', 'success')
    return redirect(url_for('admin.servicio_admin_detalle', sid=sid))
